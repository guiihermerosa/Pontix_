# -*- coding: utf-8 -*-


import csv
import logging
import os
import threading
import time
import webbrowser
from collections import defaultdict, OrderedDict
from datetime import date, datetime

from flask import (
    Flask, render_template, request, redirect, url_for,
    send_from_directory, flash, jsonify, abort
)
from werkzeug.utils import secure_filename

import config_store
import auditoria
import cadastro_local
import notificacoes
import ponto_core
from ponto_core import DeviceError

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGOS_DIR = os.path.join(BASE_DIR, "logos")
if not os.path.isdir(LOGOS_DIR):
    os.makedirs(LOGOS_DIR)

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config.update(
    SECRET_KEY="ponto-web-local-nao-precisa-ser-secreto",
    TEMPLATES_AUTO_RELOAD=True,
    JSON_SORT_KEYS=False,
)
app.logger.setLevel(logging.INFO)
notificacoes.iniciar_em_segundo_plano()


@app.route("/logos/<path:nome_arquivo>")
def logo_arquivo(nome_arquivo):
    return send_from_directory(LOGOS_DIR, nome_arquivo)


@app.context_processor
def inject_globals():
    return {
        "app_name": "Gerenciador de Ponto",
        "configuracao": config_atual(),
        "hoje_iso": date.today().isoformat(),
        "hoje_br": date.today().strftime("%d/%m/%Y"),
        "fmt_data": _data_iso_para_display,
        "fmt_hora": _hora_para_display,
        "fmt_data_hora": _data_hora_para_display,
    }


@app.errorhandler(404)
def page_not_found(error):
    return render_template("errors/404.html", error=error), 404

# progresso da ultima geracao (pra barra de progresso simples via polling)
_progresso = {"em_andamento": False, "feito": 0, "total": 0, "erro": None, "arquivo": None}
_lock = threading.Lock()


def config_atual():
    return config_store.carregar()


def _bool_form(valor):
    return valor == "on"


def _enriquecer_funcionarios_com_extras(funcionarios):
    enriquecidos = []
    for item in funcionarios or []:
        extras = cadastro_local.obter_resumo(item.get("enrollid"))
        copia = dict(item)
        copia["department"] = _departamento_efetivo(item.get("enrollid"), item.get("department", ""), extras)
        copia["cpf"] = extras.get("cpf", "")
        copia["pis"] = extras.get("pis", "")
        copia["ctps"] = extras.get("ctps", "")
        copia["data_admissao"] = extras.get("data_admissao", "")
        copia["qtd_atestados"] = extras.get("qtd_atestados", 0)
        copia["qtd_afastamentos"] = extras.get("qtd_afastamentos", 0)
        enriquecidos.append(copia)
    return enriquecidos


def _departamento_local(enrollid):
    try:
        return str(cadastro_local.obter_resumo(enrollid).get("department", "") or "").strip()
    except Exception:
        return ""


def _departamento_efetivo(enrollid, departamento="", extras=None):
    departamento = str(departamento or "").strip()
    if departamento:
        return departamento
    if extras is not None:
        departamento = str(extras.get("department", "") or "").strip()
        if departamento:
            return departamento
    return _departamento_local(enrollid)


def _int_form(valor, padrao=0):
    try:
        return int(valor)
    except (TypeError, ValueError):
        return padrao


def _cpf_form(valor):
    try:
        return cadastro_local.normalizar_cpf(valor)
    except ValueError as e:
        raise ValueError(str(e))


def _data_iso_para_display(valor):
    if not valor:
        return ""
    try:
        return datetime.strptime(valor, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return valor


def _hora_para_display(valor):
    texto = str(valor or "").strip()
    if not texto:
        return ""
    for formato in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(texto, formato).strftime("%H:%M")
        except ValueError:
            pass
    if len(texto) >= 5:
        return texto[:5]
    return texto


def _data_hora_para_display(valor):
    texto = str(valor or "").strip()
    if not texto:
        return ""
    candidatos = [
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]
    for formato in candidatos:
        try:
            return datetime.strptime(texto, formato).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            continue
    if "T" in texto:
        data_txt, hora_txt = texto.split("T", 1)
        return "{} {}".format(_data_iso_para_display(data_txt), _hora_para_display(hora_txt))
    return _data_iso_para_display(texto)


def _data_form(valor):
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except Exception:
        return None


def _texto_multilinha_para_feriados(texto):
    feriados = []
    for linha in (texto or "").splitlines():
        linha = linha.strip()
        if not linha:
            continue
        partes = [p.strip() for p in linha.split("|")]
        while len(partes) < 3:
            partes.append("")
        data_txt, nome, tipo = partes[:3]
        if not data_txt:
            continue
        feriados.append({
            "data": data_txt,
            "nome": nome or data_txt,
            "tipo": tipo or "personalizado",
            "ativo": True,
        })
    return feriados


def _texto_multilinha_para_escalas(texto):
    escalas = []
    for linha in (texto or "").splitlines():
        linha = linha.strip()
        if not linha:
            continue
        partes = [p.strip() for p in linha.split("|")]
        while len(partes) < 3:
            partes.append("")
        nome, descricao, funcionarios_txt = partes[:3]
        funcionarios = []
        for item in funcionarios_txt.split(","):
            item = item.strip()
            if item.isdigit():
                funcionarios.append(int(item))
        if nome:
            escalas.append({
                "nome": nome,
                "descricao": descricao,
                "funcionarios": funcionarios,
            })
    return escalas


def _feriados_para_texto(feriados):
    linhas = []
    for item in feriados or []:
        linhas.append("{}|{}|{}".format(item.get("data", ""), item.get("nome", ""), item.get("tipo", "personalizado")))
    return "\n".join(linhas)


def _escalas_para_texto(escalas):
    linhas = []
    for item in escalas or []:
        funcionarios = ",".join(str(f) for f in item.get("funcionarios", []) or [])
        linhas.append("{}|{}|{}".format(item.get("nome", ""), item.get("descricao", ""), funcionarios))
    return "\n".join(linhas)


def _normalizar_turno_para_view(turno, indice):
    turno = turno or {}
    sections = []
    for secao in turno.get("sections", []) or []:
        start = str(secao.get("start", "00:00")).strip() or "00:00"
        end = str(secao.get("end", "00:00")).strip() or "00:00"
        try:
            tipo = int(secao.get("type", 0) or 0)
        except (TypeError, ValueError):
            tipo = 0
        sections.append({
            "start": start,
            "end": end,
            "type": tipo,
            "is_vazia": start == "00:00" and end == "00:00",
        })

    while len(sections) < 3:
        sections.append({"start": "00:00", "end": "00:00", "type": 0, "is_vazia": True})

    sections = sections[:3]
    faixas_ativas = sum(1 for secao in sections if not secao["is_vazia"])
    nome = str(turno.get("name", "") or "").strip()
    return {
        "name": nome or "shift{}".format(indice),
        "display_name": nome or "Turno {}".format(indice),
        "cutoftime": str(turno.get("cutoftime", "00:00") or "00:00").strip() or "00:00",
        "sections": sections,
        "faixas_ativas": faixas_ativas,
    }


def _descricao_evento_marcacao(item):
    try:
        inout = int(item.get("inout", "") or item.get("event", "") or -1)
    except (TypeError, ValueError):
        inout = -1
    if inout == 1:
        return "Entrada"
    if inout == 0:
        return "Saida"

    try:
        event = int(item.get("event", 0) or 0)
    except (TypeError, ValueError):
        event = -1
    if event == 1:
        return "Entrada"
    if event == 2:
        return "Saida"
    return "Registro"


def _url_absoluta_leitor(config, caminho):
    caminho = str(caminho or "").strip()
    if not caminho:
        return ""
    if caminho.startswith("http://") or caminho.startswith("https://"):
        return caminho
    if not caminho.startswith("/"):
        caminho = "/" + caminho
    return "http://{}:{}{}".format(config.get("device_ip", ""), config.get("device_port", 80), caminho)


def _normalizar_marcacao_dashboard(item, config):
    nome = str(item.get("name", "") or "").strip()
    matricula = str(item.get("enrollid", "") or "").strip()
    departamento = _departamento_efetivo(item.get("enrollid", ""), item.get("department", ""))
    hora = str(item.get("hora", "") or "").strip()
    if not hora and item.get("time"):
        hora = str(item.get("time", "")).strip()
    acao = str(item.get("acao", "") or "").strip()
    if acao not in ("Entrada", "Saida"):
        acao = _descricao_evento_marcacao(item)
    return {
        "signature": str(item.get("signature", "") or ponto_core._assinatura_log_bruto(item)),
        "nome": nome or ("Funcionario {}".format(matricula) if matricula else "Funcionario"),
        "matricula": matricula,
        "departamento": departamento,
        "hora": hora,
        "acao": acao,
        "mode": item.get("mode", ""),
        "inout": item.get("inout", ""),
        "event": item.get("event", ""),
        "source": item.get("source", ""),
        "note": item.get("note", ""),
        "photo_url": _url_absoluta_leitor(config, item.get("photourl", "")),
    }


def _ordenar_marcacoes_desc(registros):
    return sorted(
        registros or [],
        key=lambda item: (
            str(item.get("data", "")),
            str(item.get("hora", "")),
            str(item.get("enrollid", "")),
            str(item.get("source", "")),
        ),
        reverse=True,
    )


def _dashboard_dados(config, limite=12):
    hoje = date.today()
    historico = ponto_core.ler_historico_local(config, inicio=hoje, fim=hoje, limite=2000)
    historico = _ordenar_marcacoes_desc(historico)
    recentes = [_normalizar_marcacao_dashboard(item, config) for item in historico[:limite]]
    latest = recentes[0] if recentes else None
    entradas = len([item for item in historico if str(item.get("acao", "") or "").strip() == "Entrada"])
    saidas = len([item for item in historico if str(item.get("acao", "") or "").strip() == "Saida"])
    pendentes = 0
    por_funcionario = {}
    for item in historico:
        chave = str(item.get("enrollid", "") or "").strip()
        if not chave:
            continue
        por_funcionario.setdefault(chave, []).append(item)
    for itens in por_funcionario.values():
        if not itens:
            continue
        ultimo = sorted(itens, key=lambda item: (str(item.get("hora", "")), str(item.get("signature", ""))))[-1]
        if str(ultimo.get("acao", "") or "").strip() == "Entrada":
            pendentes += 1
    try:
        total_funcionarios = len(ponto_core.get_user_ids(config))
    except Exception:
        total_funcionarios = 0
    estado = ponto_core.ler_estado_coletor(config)
    ultima_coleta = str(estado.get("ultima_coleta_em", "") or estado.get("ultima_sincronizacao", "") or "").strip()
    status_leitor = "offline"
    alerta_status = "Sem coleta recente."
    try:
        intervalo = int(config.get("intervalo_notificacao_segundos", 60) or 60)
    except (TypeError, ValueError):
        intervalo = 60
    if ultima_coleta:
        try:
            texto = ultima_coleta.strip()
            if len(texto) >= 19 and "T" in texto:
                ultima_dt = datetime.strptime(texto[:19], "%Y-%m-%dT%H:%M:%S")
            elif len(texto) >= 16 and " " in texto:
                ultima_dt = datetime.strptime(texto[:16], "%Y-%m-%d %H:%M")
            elif len(texto) == 10:
                if texto == date.today().isoformat():
                    ultima_dt = datetime.combine(date.today(), datetime.now().time())
                else:
                    ultima_dt = datetime.strptime(texto, "%Y-%m-%d")
            else:
                ultima_dt = datetime.strptime(texto[:10], "%Y-%m-%d")
            minutos = max(0, int((datetime.now() - ultima_dt).total_seconds() // 60))
            if minutos <= max(5, int((intervalo * 3) // 60) if intervalo else 5):
                status_leitor = "online"
                alerta_status = "Coleta recente há {} min.".format(minutos)
            else:
                alerta_status = "Última coleta há {} min.".format(minutos)
        except Exception:
            if ultima_coleta == date.today().isoformat():
                status_leitor = "online"
                alerta_status = "Coleta realizada hoje."
            else:
                alerta_status = "Última coleta registrada em {}".format(ultima_coleta)
    inconsistencias = len([item for item in historico if str(item.get("ajuste_manual", "")).lower() in ("true", "1")])
    return {
        "recentes": recentes,
        "latest": latest,
        "stats": {
            "total_funcionarios": total_funcionarios,
            "registros_hoje": len(historico),
            "entradas": entradas,
            "saidas": saidas,
            "pendentes": pendentes,
            "status_leitor": status_leitor,
            "ultima_coleta_em": ultima_coleta,
            "alerta_status": alerta_status,
            "ajustes_hoje": inconsistencias,
        },
    }


def _parse_data_filtro(valor, padrao):
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except Exception:
        return padrao


def _texto_filtro(valor):
    if valor is None:
        return ""
    return str(valor).strip().lower()


def _carregar_registros_operacionais(config, inicio, fim, consulta="", departamento="", turno=""):
    registros = ponto_core.ler_historico_local(config, inicio=inicio, fim=fim, limite=5000)
    consulta = _texto_filtro(consulta)
    departamento = _texto_filtro(departamento)
    turno = _texto_filtro(turno)

    filtrados = []
    for item in registros:
        nome = _texto_filtro(item.get("name", ""))
        matricula = _texto_filtro(item.get("enrollid", ""))
        depto = _texto_filtro(_departamento_efetivo(item.get("enrollid", ""), item.get("department", "")))
        shift = _texto_filtro(item.get("shift_name", ""))
        if consulta and consulta not in nome and consulta not in matricula and consulta not in depto:
            continue
        if departamento and departamento not in depto:
            continue
        if turno and turno not in shift:
            continue
        filtrados.append(item)

    agrupados = OrderedDict()
    for item in sorted(filtrados, key=lambda row: (str(row.get("data", "")), str(row.get("hora", "")), str(row.get("enrollid", "")))):
        chave = str(item.get("enrollid", "") or "").strip() or str(item.get("signature", "") or "")
        departamento_item = _departamento_efetivo(item.get("enrollid", ""), item.get("department", ""))
        grupo = agrupados.setdefault(chave, {
            "enrollid": str(item.get("enrollid", "") or "").strip(),
            "name": item.get("name", "") or "",
            "department": departamento_item,
            "shift_name": item.get("shift_name", "") or "",
            "timeline": [],
        })
        if not grupo["name"]:
            grupo["name"] = item.get("name", "") or ""
        if not grupo["department"]:
            grupo["department"] = departamento_item
        if not grupo["shift_name"]:
            grupo["shift_name"] = item.get("shift_name", "") or ""
        grupo["timeline"].append(item)

    resumo_departamentos = defaultdict(lambda: {"total": 0, "entradas": 0, "saidas": 0})
    resumo_funcionarios = []
    total_entradas = 0
    total_saidas = 0
    total_pendentes = 0
    for grupo in agrupados.values():
        items = grupo["timeline"]
        entradas = len([i for i in items if str(i.get("acao", "") or "").strip() == "Entrada"])
        saidas = len([i for i in items if str(i.get("acao", "") or "").strip() == "Saida"])
        pendente = bool(items and str(items[-1].get("acao", "") or "").strip() == "Entrada")
        total_entradas += entradas
        total_saidas += saidas
        if pendente:
            total_pendentes += 1
        depto_key = grupo.get("department") or "Sem departamento"
        resumo_departamentos[depto_key]["total"] += len(items)
        resumo_departamentos[depto_key]["entradas"] += entradas
        resumo_departamentos[depto_key]["saidas"] += saidas
        resumo_funcionarios.append({
            "enrollid": grupo.get("enrollid", ""),
            "name": grupo.get("name", ""),
            "department": grupo.get("department", "") or "Sem departamento",
            "shift_name": grupo.get("shift_name", "") or "Sem turno",
            "total": len(items),
            "entradas": entradas,
            "saidas": saidas,
            "pendente": pendente,
            "items": items,
        })

    resumo_departamentos_lista = []
    for depto, dados in sorted(resumo_departamentos.items(), key=lambda kv: kv[0].lower()):
        resumo_departamentos_lista.append({
            "name": depto,
            "total": dados["total"],
            "entradas": dados["entradas"],
            "saidas": dados["saidas"],
        })

    return {
        "registros": filtrados,
        "grupos": resumo_funcionarios,
        "departamentos": resumo_departamentos_lista,
        "total_registros": len(filtrados),
        "total_entradas": total_entradas,
        "total_saidas": total_saidas,
        "total_pendentes": total_pendentes,
    }


def _minutos_entre_horas(inicio_hora, fim_hora):
    try:
        ini = datetime.strptime(str(inicio_hora), "%H:%M:%S")
    except Exception:
        try:
            ini = datetime.strptime(str(inicio_hora), "%H:%M")
        except Exception:
            return ""
    try:
        fim = datetime.strptime(str(fim_hora), "%H:%M:%S")
    except Exception:
        try:
            fim = datetime.strptime(str(fim_hora), "%H:%M")
        except Exception:
            return ""
    delta = fim - ini
    if delta.total_seconds() < 0:
        delta += timedelta(days=1)
    minutos = int(delta.total_seconds() // 60)
    horas = minutos // 60
    mins = minutos % 60
    return "{:02d}:{:02d}".format(horas, mins)


def _espelho_por_funcionario(grupo, config):
    itens = list(grupo.get("timeline", []) or [])
    dias = OrderedDict()
    for item in sorted(itens, key=lambda row: (str(row.get("data", "")), str(row.get("hora", "")), str(row.get("signature", "")))):
        chave = str(item.get("data", "") or "").strip() or date.today().isoformat()
        dias.setdefault(chave, []).append(item)

    cpf = ""
    admissao = ""
    try:
        extras = cadastro_local.obter_resumo(grupo.get("enrollid", ""))
        cpf = extras.get("cpf", "")
        admissao = extras.get("data_admissao", "")
    except Exception:
        pass

    linhas = []
    for data_iso, registros in dias.items():
        horarios = [fmt_hora(r.get("hora", "")) for r in registros if r.get("hora")]
        horarios = [item for item in horarios if item]
        marcas = " ".join(horarios) if horarios else "--"
        jornada = marcas
        duracao = ""
        if len(horarios) >= 2:
            duracao = _minutos_entre_horas(horarios[0], horarios[-1])
        if not duracao:
            duracao = "--:--"
        tratamento = "I" if len(horarios) == 1 else "D/P" if len(horarios) > 4 else "OK"
        linhas.append([
            fmt_data(data_iso) or data_iso,
            marcas,
            jornada,
            duracao,
            grupo.get("shift_name", "") or "--",
            tratamento,
        ])

    return {
        "enrollid": grupo.get("enrollid", ""),
        "name": grupo.get("name", ""),
        "department": _departamento_efetivo(grupo.get("enrollid", ""), grupo.get("department", "")),
        "shift_name": grupo.get("shift_name", ""),
        "cpf": cpf,
        "admissao": admissao,
        "linhas": linhas,
    }


def _identificacao_relatorio(config):
    return {
        "empresa": config.get("company_name", "") or "Empresa",
        "documento": config.get("company_document", "") or "",
        "local": config.get("workplace_address", "") or "",
        "rep": config.get("rep_identifier", "") or "",
        "vendor": config.get("vendor_name", "") or "GRB Tecnologia",
    }


def _linhas_espelho_exportacao(espelhos):
    cabecalhos = [
        "Funcionario",
        "Matricula",
        "CPF",
        "Admissao",
        "Departamento",
        "Turno",
        "Data",
        "Marcacoes",
        "Jornada Prevista",
        "Total Dia",
        "Ocorrencia",
    ]
    linhas = []
    for espelho in espelhos:
        base = [
            espelho.get("name", ""),
            espelho.get("enrollid", ""),
            espelho.get("cpf", ""),
            _data_iso_para_display(espelho.get("admissao", "")),
            espelho.get("department", ""),
            espelho.get("shift_name", ""),
        ]
        registros = espelho.get("linhas", []) or []
        if not registros:
            linhas.append(base + ["Sem registros", "", "", "", ""])
            continue
        for registro in registros:
            data_txt, marcas, jornada, total_dia, _turno, ocorrencia = (list(registro) + [""] * 6)[:6]
            linhas.append(base + [data_txt, marcas, jornada, total_dia, ocorrencia])
            base = ["", "", "", "", "", ""]
    return cabecalhos, linhas


def _estilo_cabecalho_planilha(ws, titulo, subtitulo, metadados, n_colunas):
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.sheet_properties.tabColor = "1F4E78"
    ws.append([titulo])
    ws.append([subtitulo])
    ws.append([metadados])
    ws.append([])
    ws.append([])
    header_row = ws.max_row
    for row_idx in (1, 2, 3):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_colunas)
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws["A2"].font = Font(size=11, italic=True, color="4A5568")
    ws["A3"].font = Font(size=9, bold=True, color="718096")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws["A3"].alignment = Alignment(horizontal="left")
    for cel in ws[header_row]:
        cel.font = Font(bold=True, color="1F2937")
        cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws["A{}".format(header_row + 1)]
    ws.sheet_view.showGridLines = False
    return header_row


def _aplicar_faixa_linha(ws, row_number, n_colunas, destaque=False, alternada=False):
    from openpyxl.styles import Font, PatternFill, Alignment

    fill = "DCEBFA" if destaque else ("F7FAFC" if alternada else "EEF3F8")
    if destaque:
        fill = "1F4E78"
    for col_idx in range(1, n_colunas + 1):
        cel = ws.cell(row=row_number, column=col_idx)
        cel.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cel.alignment = Alignment(horizontal="center" if col_idx > 6 else "left", vertical="center", wrap_text=True)
        if destaque:
            cel.font = Font(bold=True, color="FFFFFF")
        elif col_idx == 1 or col_idx == 2:
            cel.font = Font(bold=True, color="1F2937")


def _salvar_espelho_conferencia_xlsx(caminho, config, inicio, fim, espelhos, resumo_grupos):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ident = _identificacao_relatorio(config)
    periodo_txt = "{} a {}".format(_data_iso_para_display(inicio.isoformat()), _data_iso_para_display(fim.isoformat()))
    emissao_txt = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = Workbook()

    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Espelho de Ponto Eletronico"])
    ws.append(["Empresa: {}{}".format(ident["empresa"], " | CNPJ/CPF: {}".format(ident["documento"]) if ident["documento"] else "")])
    ws.append(["Local: {}{}".format(ident["local"], " | REP: {}".format(ident["rep"]) if ident["rep"] else "")])
    ws.append(["Periodo: {} | Emissao: {}".format(periodo_txt, emissao_txt)])
    ws.append([])
    resumo_cols = [
        "Matricula",
        "Nome",
        "Departamento",
        "Turno",
        "CPF",
        "Admissao",
        "Registros",
        "Pendencia",
    ]
    ws.append(resumo_cols)
    header_row = ws.max_row
    for row_idx in (1, 2, 3, 4):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(resumo_cols))
    ws["A1"].font = Font(size=15, bold=True, color="1F4E78")
    ws["A2"].font = Font(size=10, italic=True, color="4A5568")
    ws["A3"].font = Font(size=10, italic=True, color="4A5568")
    ws["A4"].font = Font(size=9, bold=True, color="718096")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws["A3"].alignment = Alignment(horizontal="left")
    ws["A4"].alignment = Alignment(horizontal="left")
    for cel in ws[header_row]:
        cel.font = Font(bold=True, color="1F2937")
        cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, grupo in enumerate(resumo_grupos, start=1):
        ws.append([
            grupo.get("enrollid", ""),
            grupo.get("name", ""),
            grupo.get("department", ""),
            grupo.get("shift_name", ""),
            grupo.get("cpf", ""),
            _data_iso_para_display(grupo.get("admissao", "")),
            grupo.get("total", 0),
            "Sim" if grupo.get("pendente") else "Nao",
        ])
        row = ws.max_row
        _aplicar_faixa_linha(ws, row, len(resumo_cols), alternada=(idx % 2 == 0))
    ws.freeze_panes = ws["A{}".format(header_row + 1)]
    ws.auto_filter.ref = "A{}:{}{}".format(header_row, get_column_letter(len(resumo_cols)), ws.max_row)
    for i, largura in enumerate([12, 24, 20, 16, 18, 14, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws2 = wb.create_sheet("Espelho")
    espelho_cols = [
        "Funcionario",
        "Matricula",
        "CPF",
        "Admissao",
        "Departamento",
        "Turno",
        "Data",
        "Marcacoes",
        "Jornada Prevista",
        "Total Dia",
        "Ocorrencia",
    ]
    linha_atual = 1
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "0052A3"
    for indice, espelho in enumerate(espelhos, start=1):
        titulo = "{} - Matr. {}".format(espelho.get("name", "") or "Funcionario", espelho.get("enrollid", "") or "-")
        subtitulo = "{} | {} | {}".format(
            espelho.get("department", "") or "Sem departamento",
            espelho.get("shift_name", "") or "Sem turno",
            _data_iso_para_display(espelho.get("admissao", "")) or "Sem admissao",
        )
        ws2.append([titulo] + [""] * (len(espelho_cols) - 1))
        ws2.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(espelho_cols))
        cel = ws2.cell(row=linha_atual, column=1)
        cel.font = Font(bold=True, size=11, color="FFFFFF")
        cel.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cel.alignment = Alignment(horizontal="left", vertical="center")
        linha_atual += 1
        ws2.append([subtitulo] + [""] * (len(espelho_cols) - 1))
        ws2.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(espelho_cols))
        cel = ws2.cell(row=linha_atual, column=1)
        cel.font = Font(italic=True, size=9, color="1F2937")
        cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cel.alignment = Alignment(horizontal="left", vertical="center")
        linha_atual += 1
        ws2.append(espelho_cols)
        header_row_espelho = linha_atual
        for cel in ws2[header_row_espelho]:
            cel.font = Font(bold=True, color="1F2937")
            cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        linha_atual += 1
        linhas = espelho.get("linhas", []) or []
        if not linhas:
            ws2.append([
                espelho.get("name", ""),
                espelho.get("enrollid", ""),
                espelho.get("cpf", ""),
                _data_iso_para_display(espelho.get("admissao", "")),
                espelho.get("department", ""),
                espelho.get("shift_name", ""),
                "Sem registros",
                "",
                "",
                "",
                "",
            ])
            _aplicar_faixa_linha(ws2, linha_atual, len(espelho_cols), alternada=False)
            linha_atual += 1
        else:
            for idx, registro in enumerate(linhas, start=1):
                data_txt, marcas, jornada, total_dia, _turno, ocorrencia = (list(registro) + [""] * 6)[:6]
                ws2.append([
                    espelho.get("name", "") if idx == 1 else "",
                    espelho.get("enrollid", "") if idx == 1 else "",
                    espelho.get("cpf", "") if idx == 1 else "",
                    _data_iso_para_display(espelho.get("admissao", "")) if idx == 1 else "",
                    espelho.get("department", "") if idx == 1 else "",
                    espelho.get("shift_name", "") if idx == 1 else "",
                    data_txt,
                    marcas,
                    jornada,
                    total_dia,
                    ocorrencia,
                ])
                _aplicar_faixa_linha(ws2, linha_atual, len(espelho_cols), alternada=(idx % 2 == 0))
                linha_atual += 1
        linha_atual += 1
    for i, largura in enumerate([22, 12, 16, 14, 20, 16, 14, 26, 18, 12, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = largura
    ws2.freeze_panes = "A6"

    wb.save(caminho)


def _salvar_espelho_conferencia_pdf(caminho, config, inicio, fim, espelhos):
    ident = _identificacao_relatorio(config)
    periodo_txt = "{} a {}".format(_data_iso_para_display(inicio.isoformat()), _data_iso_para_display(fim.isoformat()))
    subtitulo = "Empresa: {} | Periodo: {}".format(ident["empresa"], periodo_txt)
    if ident["documento"]:
        subtitulo += " | Documento: {}".format(ident["documento"])
    if ident["rep"]:
        subtitulo += " | REP: {}".format(ident["rep"])
    if ident["local"]:
        subtitulo += " | Local: {}".format(ident["local"])

    cabecalhos = [
        "Funcionario",
        "Matricula",
        "CPF",
        "Admissao",
        "Departamento",
        "Turno",
        "Data",
        "Marcacoes",
        "Jornada Prevista",
        "Total Dia",
        "Ocorrencia",
    ]
    linhas = []
    for espelho in espelhos:
        linhas.append([
            espelho.get("name", ""),
            espelho.get("enrollid", ""),
            espelho.get("cpf", ""),
            _data_iso_para_display(espelho.get("admissao", "")),
            espelho.get("department", ""),
            espelho.get("shift_name", ""),
            "",
            "",
            "",
            "",
            "",
        ])
        linhas_espelho = espelho.get("linhas", []) or []
        if not linhas_espelho:
            linhas.append(["", "", "", "", "", "", "Sem registros", "", "", "", ""])
            continue
        for registro in linhas_espelho:
            data_txt, marcas, jornada, total_dia, _turno, ocorrencia = (list(registro) + [""] * 6)[:6]
            linhas.append(["", "", "", "", "", "", data_txt, marcas, jornada, total_dia, ocorrencia])

    ponto_core._gerar_pdf_texto(
        caminho,
        "Espelho de Ponto Eletronico",
        subtitulo,
        cabecalhos,
        linhas,
    )


def _iniciar_geracao_relatorio(inicio, fim, nome_periodo, permitir_vazio=False, nota_vazio=None, filtros=None):
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de gerar um relatorio.", "erro")
        return redirect(url_for("configuracoes"))

    with _lock:
        if _progresso["em_andamento"]:
            flash("Ja existe uma geracao de relatorio em andamento. Aguarde terminar.", "erro")
            return redirect(url_for("index"))
        _progresso.update(em_andamento=True, feito=0, total=0, erro=None, arquivo=None)

    def rodar():
        config_local = os.path.join(BASE_DIR, config["output_dir"])
        cfg = dict(config)
        cfg["output_dir"] = config_local
        if cfg.get("logo_path"):
            cfg["logo_path"] = os.path.join(LOGOS_DIR, cfg["logo_path"])

        def progresso_cb(feito, total):
            with _lock:
                _progresso.update(feito=feito, total=total)

        try:
            caminho, nome_arquivo, qtd = ponto_core.gerar_relatorio_completo(
                cfg, inicio, fim, nome_periodo, progress_cb=progresso_cb,
                permitir_vazio=permitir_vazio, nota_vazio=nota_vazio, filtros=filtros
            )
            with _lock:
                _progresso.update(em_andamento=False, arquivo=nome_arquivo, erro=None)
            auditoria.registrar_evento("geracao_relatorio", "Relatorio {} gerado com {} funcionarios".format(nome_arquivo, qtd))
        except DeviceError as e:
            with _lock:
                _progresso.update(em_andamento=False, erro=str(e))
        except Exception as e:
            with _lock:
                _progresso.update(em_andamento=False, erro="Erro inesperado: {}".format(e))

    threading.Thread(target=rodar, daemon=True).start()
    return redirect(url_for("gerando"))


def _listar_relatorios_gerados(config):
    output_dir = os.path.join(BASE_DIR, config["output_dir"])
    relatorios = []
    if os.path.isdir(output_dir):
        for nome in sorted(os.listdir(output_dir), reverse=True):
            if nome.lower().endswith((".xlsx", ".csv", ".pdf")):
                caminho = os.path.join(output_dir, nome)
                relatorios.append({
                    "nome": nome,
                    "formato": os.path.splitext(nome)[1].lstrip(".").upper(),
                    "tamanho_kb": round(os.path.getsize(caminho) / 1024, 1),
                    "gerado_em": time.strftime(
                        "%d/%m/%Y %H:%M", time.localtime(os.path.getmtime(caminho))
                    ),
                })
    return relatorios


@app.route("/")
def index():
    config = config_atual()
    relatorios = _listar_relatorios_gerados(config)
    historico_local = ponto_core.ler_historico_local(config, limite=20)
    historico_csv = ponto_core.caminho_historico_csv(config["output_dir"])
    hoje = date.today()
    historico_hoje_local = ponto_core.ler_historico_local(config, inicio=hoje, fim=hoje)
    presentes_locais = len({item.get("enrollid") for item in historico_hoje_local if item.get("enrollid")})
    movimentos_locais = len(historico_hoje_local)
    historico_dashboard = [item for item in historico_hoje_local if str(item.get("mode", "")).strip() == "8"]
    if presentes_locais or movimentos_locais:
        presentes = presentes_locais
        movimentos = movimentos_locais
    else:
        presentes = len({item.get("enrollid") for item in historico_dashboard if item.get("enrollid")})
        movimentos = len(historico_dashboard)
    try:
        total_funcionarios = len(ponto_core.get_user_ids(config))
    except Exception:
        total_funcionarios = 0
    dashboard = {
        "total_funcionarios": total_funcionarios,
        "presentes": presentes,
        "ausentes": max(0, total_funcionarios - presentes),
        "movimentos": movimentos,
    }
    return render_template(
        "index.html",
        config=config,
        configurado=config_store.esta_configurado(config),
        relatorios=relatorios,
        historico_local=historico_local,
        historico_csv_disponivel=os.path.isfile(historico_csv),
        dashboard=dashboard,
        hoje=hoje.isoformat(),
    )


@app.route("/relatorios")
def relatorios():
    config = config_atual()
    filtros = _filtros_relatorio_request(request.args)
    return render_template(
        "relatorios.html",
        config=config,
        configurado=config_store.esta_configurado(config),
        relatorios=_listar_relatorios_gerados(config),
        filtros=filtros,
        resumo_filtros=_resumo_filtros(filtros),
        opcoes_relatorio=_opcoes_relatorio(config),
        hoje=date.today().isoformat(),
    )


@app.route("/dashboard")
def dashboard():
    config = config_atual()
    dados = _dashboard_dados(config)
    return render_template(
        "dashboard.html",
        config=config,
        configurado=config_store.esta_configurado(config),
        dashboard_stats=dados["stats"],
        dashboard_recentes=dados["recentes"],
        dashboard_latest=dados["latest"],
        hoje=date.today().isoformat(),
    )


@app.route("/dashboard/api")
def dashboard_api():
    config = config_atual()
    dados = _dashboard_dados(config)
    return jsonify({
        "ok": True,
        "stats": dados["stats"],
        "recentes": dados["recentes"],
        "latest": dados["latest"],
    })


def _operacional_contexto(config, args):
    hoje = date.today()
    data_txt = args.get("data", hoje.isoformat())
    inicio = _parse_data_filtro(args.get("inicio", ""), _parse_data_filtro(data_txt, hoje))
    fim = _parse_data_filtro(args.get("fim", ""), inicio)
    if fim < inicio:
        fim = inicio
    consulta = args.get("q", "").strip()
    departamento = args.get("departamento", "").strip()
    turno = args.get("turno", "").strip()
    dados = _carregar_registros_operacionais(config, inicio, fim, consulta=consulta, departamento=departamento, turno=turno)
    filtros = {
        "inicio": inicio.isoformat(),
        "fim": fim.isoformat(),
        "q": consulta,
        "departamento": departamento,
        "turno": turno,
    }
    return dados, filtros


def _texto_busca(valor):
    texto = str(valor or "").strip()
    if not texto:
        return ""
    return texto


def _filtros_relatorio_request(source):
    source = source or {}
    return {
        "matricula": _texto_busca(source.get("matricula", "")),
        "nome": _texto_busca(source.get("nome", "")),
        "departamento": _texto_busca(source.get("departamento", "")),
        "turno": _texto_busca(source.get("turno", "")),
        "periodo_inicio": _texto_busca(source.get("periodo_inicio", source.get("data_inicial", ""))),
        "periodo_fim": _texto_busca(source.get("periodo_fim", source.get("data_final", ""))),
    }


def _resumo_filtros(filtros):
    itens = []
    labels = {
        "matricula": "Matrícula",
        "nome": "Nome",
        "departamento": "Departamento",
        "turno": "Turno",
        "periodo_inicio": "Período inicial",
        "periodo_fim": "Período final",
    }
    for chave, rotulo in labels.items():
        valor = str((filtros or {}).get(chave, "")).strip()
        if valor:
            itens.append({"label": rotulo, "value": valor})
    return itens


def _opcoes_relatorio(config):
    opcoes = {
        "departamentos": [],
        "turnos": [],
        "funcionarios": [],
    }
    if not config_store.esta_configurado(config):
        return opcoes
    try:
        funcionarios = _enriquecer_funcionarios_com_extras(ponto_core.listar_funcionarios(config))
    except Exception:
        funcionarios = []

    nomes = []
    matriculas = []
    departamentos = []
    turnos = []
    for funcionario in funcionarios:
        nome = str(funcionario.get("name", "") or "").strip()
        matricula = str(funcionario.get("enrollid", "") or "").strip()
        departamento = str(funcionario.get("department", "") or "").strip()
        turno = str(funcionario.get("shift_name", "") or "").strip()
        if nome:
            nomes.append(nome)
        if matricula:
            matriculas.append(matricula)
        if departamento:
            departamentos.append(departamento)
        if turno:
            turnos.append(turno)

    opcoes["funcionarios"] = sorted(set(nomes), key=lambda valor: valor.lower())
    opcoes["matriculas"] = sorted(
        set(matriculas),
        key=lambda valor: (0, int(valor)) if valor.isdigit() else (1, valor.lower()),
    )
    opcoes["departamentos"] = sorted(set(departamentos), key=lambda valor: valor.lower())
    opcoes["turnos"] = sorted(set(turnos), key=lambda valor: valor.lower())
    return opcoes


def _paginacao_total(total_itens, page, per_page):
    per_page = max(1, int(per_page or 20))
    total = int(total_itens or 0)
    total_paginas = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(int(page or 1), total_paginas))
    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "pages": total_paginas,
        "has_prev": page > 1,
        "has_next": page < total_paginas,
        "prev_page": max(1, page - 1),
        "next_page": min(total_paginas, page + 1),
    }


def _aplicar_paginacao(lista, page=1, per_page=20):
    paginacao = _paginacao_total(len(lista or []), page, per_page)
    inicio = (paginacao["page"] - 1) * paginacao["per_page"]
    fim = inicio + paginacao["per_page"]
    return (lista or [])[inicio:fim], paginacao


def _filtrar_funcionarios(lista, filtros):
    filtros = filtros or {}
    termo = _texto_busca(filtros.get("q", ""))
    departamento = _texto_busca(filtros.get("departamento", "")).lower()
    if not termo and not departamento:
        return list(lista or [])

    filtrados = []
    for item in lista or []:
        nome = str(item.get("name", "") or "")
        matricula = str(item.get("enrollid", "") or "")
        depto = str(item.get("department", "") or "")
        texto = "{} {} {}".format(matricula, nome, depto).lower()
        if termo and termo.lower() not in texto:
            continue
        if departamento and departamento not in depto.lower():
            continue
        filtrados.append(item)
    return filtrados


@app.route("/conferencia", methods=["GET"])
def conferencia():
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de abrir a conferência.", "erro")
        return redirect(url_for("configuracoes"))

    dados, filtros = _operacional_contexto(config, request.args)
    try:
        pagina = int(request.args.get("page", 1))
    except ValueError:
        pagina = 1
    try:
        por_pagina = int(request.args.get("per_page", 12))
    except ValueError:
        por_pagina = 12
    grupos_paginados, paginacao = _aplicar_paginacao(dados["grupos"], pagina, por_pagina)
    turnos = []
    try:
        turnos = ponto_core.get_shifts(config)
    except Exception:
        turnos = []

    alertas = []
    if dados["total_pendentes"]:
        alertas.append("{} funcionário(s) aguardando saída.".format(dados["total_pendentes"]))
    if not dados["registros"]:
        alertas.append("Nenhuma marcação encontrada para os filtros atuais.")

    return render_template(
        "conferencia.html",
        config=config,
        configurado=True,
        filtro=filtros,
        conferencias=grupos_paginados,
        paginacao=paginacao,
        departamentos=dados["departamentos"],
        registros=dados["registros"],
        totais={
            "registros": dados["total_registros"],
            "entradas": dados["total_entradas"],
            "saidas": dados["total_saidas"],
            "pendentes": dados["total_pendentes"],
        },
        alertas=alertas,
        turnos=turnos,
        hoje=date.today().isoformat(),
    )


@app.route("/conferencia/ajustar", methods=["POST"])
def conferencia_ajustar():
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o leitor antes de ajustar as marcações.", "erro")
        return redirect(url_for("conferencia"))

    original_signature = request.form.get("original_signature", "").strip()
    acao = request.form.get("acao", "").strip()
    observacao = request.form.get("observacao", "").strip()
    redirecionar = request.form.get("redirect_to", "") or url_for("conferencia")

    try:
        ponto_core.registrar_ajuste_local(config, original_signature, acao, observacao, ajustado_por="web")
        auditoria.registrar_evento(
            "ajuste_marcacao",
            "Ajuste de marcacao {} para {}: {}".format(original_signature, acao, observacao),
        )
        flash("Ajuste salvo com sucesso.", "ok")
    except Exception as e:
        flash(str(e), "erro")

    return redirect(redirecionar)


def _nome_arquivo_operacional(formato, inicio, fim):
    return "Operacional_{}_a_{}.{}".format(
        inicio.strftime("%Y-%m-%d"),
        fim.strftime("%Y-%m-%d"),
        formato,
    )


@app.route("/conferencia/exportar/<formato>")
def conferencia_exportar(formato):
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o leitor antes de exportar os dados.", "erro")
        return redirect(url_for("conferencia"))

    dados, filtros = _operacional_contexto(config, request.args)
    formato = str(formato or "").lower()
    if formato not in ("csv", "xlsx", "pdf"):
        abort(404)

    output_dir = os.path.join(BASE_DIR, config["output_dir"])
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir)

    inicio = _parse_data_filtro(filtros["inicio"], date.today())
    fim = _parse_data_filtro(filtros["fim"], inicio)
    nome_arquivo = _nome_arquivo_operacional(formato, inicio, fim)
    caminho = os.path.join(output_dir, nome_arquivo)

    espelhos = [_espelho_por_funcionario(grupo, config) for grupo in dados["grupos"]]
    cabecalhos_csv, linhas_csv = _linhas_espelho_exportacao(espelhos)

    if formato == "csv":
        with open(caminho, "w", newline="", encoding="utf-8-sig") as f_csv:
            writer = csv.writer(f_csv, delimiter=";")
            writer.writerow(cabecalhos_csv)
            for linha in linhas_csv:
                writer.writerow(linha)
    elif formato == "xlsx":
        _salvar_espelho_conferencia_xlsx(caminho, config, inicio, fim, espelhos, dados["grupos"])
    else:
        _salvar_espelho_conferencia_pdf(caminho, config, inicio, fim, espelhos)

    return send_from_directory(output_dir, nome_arquivo, as_attachment=True)


@app.route("/turnos", methods=["GET", "POST"])
def turnos():
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de gerenciar turnos.", "erro")
        return redirect(url_for("configuracoes"))

    if request.method == "POST":
        shifts = []
        turnos_flags = []
        for idx in range(1, 9):
            sections = []
            sections_flags = []
            for sec in range(1, 4):
                start = request.form.get("shift_{}_section_{}_start".format(idx, sec), "00:00").strip() or "00:00"
                end = request.form.get("shift_{}_section_{}_end".format(idx, sec), "00:00").strip() or "00:00"
                try:
                    type_sec = int(request.form.get("shift_{}_section_{}_type".format(idx, sec), "0"))
                except ValueError:
                    type_sec = 0
                active = request.form.get("shift_{}_section_{}_active".format(idx, sec)) == "on"
                sections.append({"start": start, "end": end, "type": type_sec})
                sections_flags.append({"active": active})

            shifts.append({
                "name": request.form.get("shift_{}_name".format(idx), "shift{}".format(idx)).strip() or "shift{}".format(idx),
                "cutoftime": request.form.get("shift_{}_cutoftime".format(idx), "00:00").strip() or "00:00",
                "sections": sections,
            })
            turnos_flags.append({"sections": sections_flags})

        try:
            ponto_core.set_shifts(config, shifts)
            ponto_core.set_holidays(config, None)
            ponto_core.set_belltimes(config, None)
            ponto_core.set_devlock(config, None)
            config["turnos_config"] = turnos_flags
            config_store.salvar(config)
            auditoria.registrar_evento("alteracao_turnos", "Turnos do leitor atualizados")
            flash("Turnos salvos no leitor.", "ok")
        except DeviceError as e:
            flash(str(e), "erro")
        except Exception as e:
            flash("Erro inesperado ao salvar turnos: {}".format(e), "erro")
        return redirect(url_for("turnos"))

    try:
        shifts = ponto_core.get_shifts(config)
        erro = None
    except DeviceError as e:
        shifts = []
        erro = str(e)

    shifts_view = ponto_core.turnos_com_config(config, shifts)

    return render_template("turnos.html", shifts=shifts_view[:8], erro=erro)


@app.route("/regras", methods=["GET", "POST"])
def regras():
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de ajustar as regras.", "erro")
        return redirect(url_for("configuracoes"))

    regras_cfg = config.get("regras_ponto", {})
    if request.method == "POST":
        regras_cfg = {
            "banco_horas": {
                "ativo": _bool_form(request.form.get("banco_ativo")),
                "limite_minutos": _int_form(request.form.get("banco_limite_minutos"), 480),
                "validade_dias": _int_form(request.form.get("banco_validade_dias"), 30),
                "compensacao_minutos": _int_form(request.form.get("banco_compensacao_minutos"), 0),
            },
            "horas_extras": {
                "ativo": _bool_form(request.form.get("extras_ativo")),
                "percentual_50": _int_form(request.form.get("extras_50"), 50),
                "percentual_70": _int_form(request.form.get("extras_70"), 70),
                "percentual_100": _int_form(request.form.get("extras_100"), 100),
                "domingos_percentual": _int_form(request.form.get("extras_domingo"), 100),
                "feriados_percentual": _int_form(request.form.get("extras_feriado"), 100),
                "noturnas_percentual": _int_form(request.form.get("extras_noturna"), 70),
            },
            "adicional_noturno": {
                "ativo": _bool_form(request.form.get("noturno_ativo")),
                "inicio": request.form.get("noturno_inicio", "22:00").strip() or "22:00",
                "fim": request.form.get("noturno_fim", "05:00").strip() or "05:00",
                "percentual": _int_form(request.form.get("noturno_percentual"), 20),
            },
            "tolerancias": {
                "entrada_minutos": _int_form(request.form.get("tol_entrada"), 5),
                "saida_minutos": _int_form(request.form.get("tol_saida"), 10),
            },
            "intervalo": {
                "minimo_minutos": _int_form(request.form.get("intervalo_minimo"), 30),
                "maximo_minutos": _int_form(request.form.get("intervalo_maximo"), 120),
                "avisar": _bool_form(request.form.get("intervalo_avisar")),
            },
            "feriados": _texto_multilinha_para_feriados(request.form.get("feriados_texto", "")),
            "escalas": _texto_multilinha_para_escalas(request.form.get("escalas_texto", "")),
        }
        config["regras_ponto"] = regras_cfg
        config_store.salvar(config)
        auditoria.registrar_evento("alteracao_regras", "Regras de ponto atualizadas")
        flash("Regras salvas.", "ok")
        return redirect(url_for("regras"))

    return render_template(
        "regras.html",
        config=config,
        regras=regras_cfg,
        feriados_texto=_feriados_para_texto(regras_cfg.get("feriados", [])),
        escalas_texto=_escalas_para_texto(regras_cfg.get("escalas", [])),
    )


@app.route("/auditoria")
def auditoria_view():
    config = config_atual()
    eventos = auditoria.ler_eventos(limite=200)
    logs_reais = []
    erro_logs = None
    if config_store.esta_configurado(config):
        try:
            hoje = date.today()
            logs_reais = ponto_core.ler_historico_local(config, inicio=hoje, fim=hoje, limite=200)
        except DeviceError as e:
            erro_logs = str(e)
    else:
        erro_logs = "Configure a senha do leitor para carregar os logs do equipamento."

    return render_template(
        "auditoria.html",
        eventos=eventos,
        logs_reais=logs_reais,
        erro_logs=erro_logs,
        config=config,
    )


@app.route("/auditoria/logs")
def auditoria_logs():
    config = config_atual()
    if not config_store.esta_configurado(config):
        return jsonify({"ok": False, "mensagem": "Configure a senha do leitor para carregar os logs do equipamento.", "logs": []})

    try:
        hoje = date.today()
        logs_reais = ponto_core.ler_historico_local(config, inicio=hoje, fim=hoje, limite=200)
        return jsonify({"ok": True, "logs": logs_reais})
    except Exception as e:
        return jsonify({"ok": False, "mensagem": "Erro inesperado: {}".format(e), "logs": []})


@app.route("/gerar", methods=["POST"])
def gerar():
    tipo = request.form.get("tipo")
    filtros = _filtros_relatorio_request(request.form)
    if tipo == "semana":
        inicio, fim = ponto_core.periodo_semana_passada()
        nome_periodo = "Semanal"
    elif tipo == "mes":
        inicio, fim = ponto_core.periodo_mes_passado()
        nome_periodo = "Mensal"
    elif tipo == "semana_atual":
        inicio, fim = ponto_core.periodo_semana_atual()
        nome_periodo = "SemanaAtual"
    elif tipo == "mes_atual":
        inicio, fim = ponto_core.periodo_mes_atual()
        nome_periodo = "MesAtual"
    elif tipo == "personalizado":
        try:
            inicio = _data_form(filtros["periodo_inicio"])
            fim = _data_form(filtros["periodo_fim"])
            if inicio is None or fim is None:
                raise ValueError
        except ValueError:
            flash("Datas invalidas. Use o seletor de data do formulario.", "erro")
            return redirect(url_for("index"))
        if fim < inicio:
            flash("A data final precisa ser depois da data inicial.", "erro")
            return redirect(url_for("index"))
        nome_periodo = "Personalizado"
    elif tipo == "validacao_inicial":
        inicio = date.today()
        fim = date.today()
        nome_periodo = "Validacao_Hoje"
        return _iniciar_geracao_relatorio(
            inicio,
            fim,
            nome_periodo,
            permitir_vazio=True,
            nota_vazio="Nenhuma marcacao facial foi encontrada hoje.",
        )
    else:
        flash("Escolha um periodo valido.", "erro")
        return redirect(url_for("index"))
    return _iniciar_geracao_relatorio(inicio, fim, nome_periodo, filtros=filtros)


@app.route("/validacao-inicial", methods=["POST"])
def validacao_inicial():
    hoje = date.today()
    return _iniciar_geracao_relatorio(
        hoje,
        hoje,
        "Validacao_Hoje",
        permitir_vazio=True,
        nota_vazio="Nenhuma marcacao facial foi encontrada hoje.",
    )


@app.route("/gerando")
def gerando():
    return render_template("gerando.html")


@app.route("/progresso")
def progresso():
    with _lock:
        return jsonify(dict(_progresso))


@app.route("/download/<path:nome_arquivo>")
def download(nome_arquivo):
    config = config_atual()
    output_dir = os.path.join(BASE_DIR, config["output_dir"])
    return send_from_directory(output_dir, nome_arquivo, as_attachment=True)


@app.route("/download-historico")
def download_historico():
    config = config_atual()
    caminho_csv = ponto_core.caminho_historico_csv(config["output_dir"])
    if not os.path.isfile(caminho_csv):
        flash("Ainda nao existe historico local salvo.", "erro")
        return redirect(url_for("index"))
    return send_from_directory(
        os.path.dirname(caminho_csv),
        os.path.basename(caminho_csv),
        as_attachment=True,
    )


@app.route("/excluir/<path:nome_arquivo>", methods=["POST"])
def excluir(nome_arquivo):
    config = config_atual()
    output_dir = os.path.join(BASE_DIR, config["output_dir"])
    caminho = os.path.join(output_dir, secure_filename(nome_arquivo))
    if os.path.isfile(caminho):
        os.remove(caminho)
        auditoria.registrar_evento("exclusao_relatorio", "Relatorio {}".format(nome_arquivo))
        flash("Relatorio excluido.", "ok")
    return redirect(url_for("index"))


@app.route("/fechamento-local", methods=["POST"])
def fechamento_local():
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de gerar um fechamento local.", "erro")
        return redirect(url_for("configuracoes"))

    tipo = request.form.get("tipo", "mes_atual")
    if tipo == "mes_passado":
        inicio, fim = ponto_core.periodo_mes_passado()
        nome_periodo = "FechamentoMesPassado"
    else:
        inicio, fim = ponto_core.periodo_mes_atual()
        nome_periodo = "FechamentoMesAtual"

    try:
        _, nome_arquivo, qtd = ponto_core.gerar_fechamento_local(config, inicio, fim, nome_periodo)
        auditoria.registrar_evento("fechamento_local", "Fechamento local {} gerado com {} registros".format(nome_arquivo, qtd))
        flash("Fechamento local gerado com {} marcacao(oes): {}".format(qtd, nome_arquivo), "ok")
    except DeviceError as e:
        flash(str(e), "erro")
    except Exception as e:
        flash("Erro inesperado ao gerar fechamento local: {}".format(e), "erro")

    return redirect(url_for("index"))


@app.route("/funcionarios")
def funcionarios():
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de gerenciar funcionarios.", "erro")
        return redirect(url_for("configuracoes"))

    filtros = {
        "q": request.args.get("q", "").strip(),
        "departamento": request.args.get("departamento", "").strip(),
    }
    try:
        pagina = int(request.args.get("page", 1))
    except ValueError:
        pagina = 1
    try:
        por_pagina = int(request.args.get("per_page", 12))
    except ValueError:
        por_pagina = 12

    try:
        lista = ponto_core.listar_funcionarios(config)
        lista = _enriquecer_funcionarios_com_extras(lista)
        lista = _filtrar_funcionarios(lista, filtros)
        proximo_id = ponto_core.get_next_enrollid(config)
        turnos = ponto_core.get_shifts(config)
        erro = None
    except DeviceError as e:
        lista = []
        proximo_id = None
        turnos = []
        erro = str(e)

    funcionarios_paginados, paginacao = _aplicar_paginacao(lista, pagina, por_pagina)

    return render_template(
        "funcionarios.html",
        funcionarios=funcionarios_paginados,
        proximo_id=proximo_id,
        erro=erro,
        turnos=turnos,
        filtro_funcionarios=filtros,
        paginacao=paginacao,
    )


@app.route("/funcionarios/novo", methods=["POST"])
def funcionarios_novo():
    config = config_atual()
    nome = request.form.get("name", "").strip()
    departamento = request.form.get("department", "").strip()
    cpf = request.form.get("cpf", "").strip()
    pis = request.form.get("pis", "").strip()
    ctps = request.form.get("ctps", "").strip()
    data_admissao = request.form.get("data_admissao", "").strip()
    enrollid_str = request.form.get("enrollid", "").strip()
    shiftid_str = request.form.get("shiftid", "0").strip()

    if not nome:
        flash("Informe o nome do funcionario.", "erro")
        return redirect(url_for("funcionarios"))

    if not enrollid_str:
        flash("Informe a matricula (ou clique em 'sugerir' antes de salvar).", "erro")
        return redirect(url_for("funcionarios"))

    try:
        enrollid = int(enrollid_str)
    except ValueError:
        flash("A matricula precisa ser um numero.", "erro")
        return redirect(url_for("funcionarios"))

    try:
        shiftid = int(shiftid_str)
    except ValueError:
        shiftid = 0

    try:
        cpf = _cpf_form(cpf)
    except ValueError as e:
        flash(str(e), "erro")
        return redirect(url_for("funcionarios"))

    try:
        ja_existe = ponto_core.check_user_id(config, enrollid)
        if ja_existe:
            flash("Essa matricula ja esta em uso. Escolha outra.", "erro")
            return redirect(url_for("funcionarios"))

        ponto_core.set_user_info(config, enrollid, nome, departamento, shiftid=shiftid)
        if cpf:
            cadastro_local.salvar_cpf(enrollid, cpf)
        cadastro_local.salvar_dados_profissionais(enrollid, pis=pis, ctps=ctps, data_admissao=data_admissao, department=departamento)
        auditoria.registrar_evento("cadastro_funcionario", "Funcionario {} cadastrado".format(enrollid))
        flash(
            "Funcionario \"{}\" cadastrado com a matricula {}. Agora falta cadastrar a face dele "
            "direto no leitor, olhando para a camera do equipamento.".format(nome, enrollid),
            "ok",
        )
    except DeviceError as e:
        flash(str(e), "erro")

    return redirect(url_for("funcionarios"))


@app.route("/funcionarios/editar/<int:enrollid>", methods=["POST"])
def funcionarios_editar(enrollid):
    config = config_atual()
    nome = request.form.get("name", "").strip()
    departamento = request.form.get("department", "").strip()
    cpf = request.form.get("cpf", "").strip()
    pis = request.form.get("pis", "").strip()
    ctps = request.form.get("ctps", "").strip()
    data_admissao = request.form.get("data_admissao", "").strip()
    shiftid_str = request.form.get("shiftid", "0").strip()

    if not nome:
        flash("Informe o nome do funcionario.", "erro")
        return redirect(url_for("funcionarios"))

    try:
        shiftid = int(shiftid_str)
    except ValueError:
        shiftid = 0

    try:
        cpf = _cpf_form(cpf)
    except ValueError as e:
        flash(str(e), "erro")
        return redirect(url_for("funcionarios"))

    try:
        ponto_core.set_user_info(config, enrollid, nome, departamento, shiftid=shiftid)
        cadastro_local.salvar_cpf(enrollid, cpf)
        cadastro_local.salvar_dados_profissionais(enrollid, pis=pis, ctps=ctps, data_admissao=data_admissao, department=departamento)
        auditoria.registrar_evento("edicao_funcionario", "Funcionario {} atualizado".format(enrollid))
        flash("Dados de \"{}\" atualizados.".format(nome), "ok")
    except DeviceError as e:
        flash(str(e), "erro")

    return redirect(url_for("funcionarios"))


@app.route("/funcionarios/proximo-id")
def funcionarios_proximo_id():
    config = config_atual()
    try:
        proximo = ponto_core.get_next_enrollid(config)
        return jsonify({"ok": True, "enrollid": proximo})
    except DeviceError as e:
        return jsonify({"ok": False, "mensagem": str(e)})


@app.route("/funcionarios/verificar-id")
def funcionarios_verificar_id():
    config = config_atual()
    enrollid = request.args.get("enrollid", "")
    if not enrollid.isdigit():
        return jsonify({"ok": False, "mensagem": "Matricula invalida."})
    try:
        existe = ponto_core.check_user_id(config, int(enrollid))
        if existe:
            return jsonify({"ok": True, "disponivel": False, "mensagem": "Matricula ja esta em uso."})
        return jsonify({"ok": True, "disponivel": True, "mensagem": "Matricula disponivel."})
    except DeviceError as e:
        return jsonify({"ok": False, "mensagem": str(e)})


@app.route("/funcionarios/<int:enrollid>/perfil")
def funcionario_perfil(enrollid):
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de acessar o perfil do funcionario.", "erro")
        return redirect(url_for("configuracoes"))

    try:
        funcionarios = ponto_core.listar_funcionarios(config)
        funcionario = next((item for item in funcionarios if int(item.get("enrollid", -1)) == enrollid), None)
        if not funcionario:
            flash("Funcionario nao encontrado no leitor.", "erro")
            return redirect(url_for("funcionarios"))
        funcionario["department"] = _departamento_efetivo(enrollid, funcionario.get("department", ""))

        extras = cadastro_local.obter_funcionario(enrollid)
        return render_template(
            "funcionario_perfil.html",
            funcionario=funcionario,
            cpf=extras.get("cpf", ""),
            pis=extras.get("pis", ""),
            ctps=extras.get("ctps", ""),
            data_admissao=extras.get("data_admissao", ""),
            atestados=extras.get("atestados", []) or [],
            afastamentos=extras.get("afastamentos", []) or [],
        )
    except DeviceError as e:
        flash(str(e), "erro")
        return redirect(url_for("funcionarios"))


@app.route("/funcionarios/<int:enrollid>/perfil", methods=["POST"])
def funcionario_perfil_salvar(enrollid):
    cpf = request.form.get("cpf", "").strip()
    pis = request.form.get("pis", "").strip()
    ctps = request.form.get("ctps", "").strip()
    data_admissao = request.form.get("data_admissao", "").strip()
    try:
        cpf = cadastro_local.normalizar_cpf(cpf)
        cadastro_local.salvar_cpf(enrollid, cpf)
        cadastro_local.salvar_dados_profissionais(enrollid, pis=pis, ctps=ctps, data_admissao=data_admissao)
        flash("CPF atualizado com sucesso.", "ok")
    except ValueError as e:
        flash(str(e), "erro")
    except Exception as e:
        flash("Nao consegui salvar o CPF: {}".format(e), "erro")
    return redirect(url_for("funcionario_perfil", enrollid=enrollid))


@app.route("/funcionarios/<int:enrollid>/atestado", methods=["POST"])
def funcionario_atestado_upload(enrollid):
    arquivo = request.files.get("arquivo")
    data_emissao = request.form.get("data_emissao", "").strip()
    validade_ate = request.form.get("validade_ate", "").strip()
    observacoes = request.form.get("observacoes", "").strip()

    try:
        cadastro_local.adicionar_atestado(
            enrollid,
            arquivo,
            data_emissao=data_emissao,
            validade_ate=validade_ate,
            observacoes=observacoes,
        )
        auditoria.registrar_evento("atestado_cadastrado", "Atestado anexado ao funcionario {}".format(enrollid))
        flash("Atestado anexado com sucesso.", "ok")
    except ValueError as e:
        flash(str(e), "erro")
    except Exception as e:
        flash("Nao consegui anexar o atestado: {}".format(e), "erro")

    return redirect(url_for("funcionario_perfil", enrollid=enrollid))


@app.route("/funcionarios/<int:enrollid>/atestado/<atestado_id>/download")
def funcionario_atestado_download(enrollid, atestado_id):
    caminho = cadastro_local.caminho_atestado(enrollid, atestado_id)
    if not caminho:
        abort(404)
    return send_from_directory(os.path.dirname(caminho), os.path.basename(caminho), as_attachment=True)


@app.route("/funcionarios/<int:enrollid>/atestado/<atestado_id>/excluir", methods=["POST"])
def funcionario_atestado_excluir(enrollid, atestado_id):
    removido = cadastro_local.excluir_atestado(enrollid, atestado_id)
    if removido:
        auditoria.registrar_evento("atestado_excluido", "Atestado removido do funcionario {}".format(enrollid))
        flash("Atestado removido.", "ok")
    else:
        flash("Atestado nao encontrado.", "erro")
    return redirect(url_for("funcionario_perfil", enrollid=enrollid))


@app.route("/funcionarios/<int:enrollid>/afastamento", methods=["POST"])
def funcionario_afastamento_upload(enrollid):
    tipo = request.form.get("tipo", "").strip()
    data_inicio = request.form.get("data_inicio", "").strip()
    data_fim = request.form.get("data_fim", "").strip()
    observacoes = request.form.get("observacoes", "").strip()
    try:
        cadastro_local.adicionar_afastamento(
            enrollid,
            tipo=tipo,
            data_inicio=data_inicio,
            data_fim=data_fim,
            observacoes=observacoes,
        )
        auditoria.registrar_evento("afastamento_cadastrado", "Afastamento anexado ao funcionario {}".format(enrollid))
        flash("Afastamento registrado com sucesso.", "ok")
    except ValueError as e:
        flash(str(e), "erro")
    except Exception as e:
        flash("Nao consegui registrar o afastamento: {}".format(e), "erro")
    return redirect(url_for("funcionario_perfil", enrollid=enrollid))


@app.route("/funcionarios/<int:enrollid>/afastamento/<afastamento_id>/excluir", methods=["POST"])
def funcionario_afastamento_excluir(enrollid, afastamento_id):
    removido = cadastro_local.excluir_afastamento(enrollid, afastamento_id)
    if removido:
        auditoria.registrar_evento("afastamento_excluido", "Afastamento removido do funcionario {}".format(enrollid))
        flash("Afastamento removido.", "ok")
    else:
        flash("Afastamento nao encontrado.", "erro")
    return redirect(url_for("funcionario_perfil", enrollid=enrollid))


@app.route("/funcionarios/<int:enrollid>/exportar/<formato>")
def funcionario_exportar(enrollid, formato):
    config = config_atual()
    if not config_store.esta_configurado(config):
        flash("Configure o IP e a senha do leitor antes de exportar os dados do funcionario.", "erro")
        return redirect(url_for("funcionario_perfil", enrollid=enrollid))

    formato = str(formato or "").lower()
    if formato not in ("xlsx", "pdf"):
        abort(404)

    try:
        funcionarios = ponto_core.listar_funcionarios(config)
        funcionario = next((item for item in funcionarios if int(item.get("enrollid", -1)) == enrollid), None)
        if not funcionario:
            flash("Funcionario nao encontrado no leitor.", "erro")
            return redirect(url_for("funcionarios"))
        funcionario["department"] = _departamento_efetivo(enrollid, funcionario.get("department", ""))
        extras = cadastro_local.obter_funcionario(enrollid)
    except DeviceError as e:
        flash(str(e), "erro")
        return redirect(url_for("funcionario_perfil", enrollid=enrollid))

    output_dir = os.path.join(BASE_DIR, config["output_dir"])
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir)

    nome_base = "Funcionario_{}_{}".format(enrollid, secure_filename(funcionario.get("name", "perfil")) or "perfil")
    nome_arquivo = "{}.{}".format(nome_base, formato)
    caminho = os.path.join(output_dir, nome_arquivo)

    resumo = [
        ["Matricula", funcionario.get("enrollid", "")],
        ["Nome", funcionario.get("name", "")],
        ["Departamento", funcionario.get("department", "")],
        ["Turno", funcionario.get("shift_name", "")],
        ["CPF", extras.get("cpf", "")],
        ["PIS", extras.get("pis", "")],
        ["CTPS", extras.get("ctps", "")],
        ["Admissao", _data_iso_para_display(extras.get("data_admissao", ""))],
        ["Atestados", len(extras.get("atestados", []) or [])],
        ["Afastamentos/Licencas", len(extras.get("afastamentos", []) or [])],
    ]

    if formato == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ident = _identificacao_relatorio(config)
        periodo_txt = "Gerado em {}".format(datetime.now().strftime("%d/%m/%Y %H:%M"))

        ws = wb.active
        ws.title = "Cadastro"
        ws.sheet_properties.tabColor = "1F4E78"
        ws.append(["Perfil do Funcionario"])
        ws.append(["Empresa: {}{}".format(ident["empresa"], " | CNPJ/CPF: {}".format(ident["documento"]) if ident["documento"] else "")])
        ws.append(["Local: {}{}".format(ident["local"], " | REP: {}".format(ident["rep"]) if ident["rep"] else "")])
        ws.append([periodo_txt])
        ws.append([])
        ws.append(["Campo", "Valor"])
        header_row = ws.max_row
        for row_idx in (1, 2, 3, 4):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
        ws["A2"].font = Font(size=10, italic=True, color="4A5568")
        ws["A3"].font = Font(size=10, italic=True, color="4A5568")
        ws["A4"].font = Font(size=9, bold=True, color="718096")
        for cel in ws[header_row]:
            cel.font = Font(bold=True, color="1F2937")
            cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for linha in resumo:
            ws.append(linha)
        for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=1):
            for cel in row:
                cel.alignment = Alignment(vertical="top")
                cel.fill = PatternFill(start_color="F7FAFC" if idx % 2 else "EEF3F8", end_color="F7FAFC" if idx % 2 else "EEF3F8", fill_type="solid")
            row[0].font = Font(bold=True, color="1F2937")
        ws.freeze_panes = ws["A{}".format(header_row + 1)]
        ws.auto_filter.ref = "A{}:B{}".format(header_row, ws.max_row)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 48

        ws2 = wb.create_sheet("Documentos")
        ws2.sheet_properties.tabColor = "0052A3"
        ws2.append(["Afastamentos"])
        ws2.append(["Funcionario: {}".format(funcionario.get("name", ""))])
        ws2.append(["Matricula: {}".format(funcionario.get("enrollid", ""))])
        ws2.append([])
        ws2.append(["Tipo", "Inicio", "Fim", "Observacoes", "Criado em"])
        header_row2 = ws2.max_row
        for row_idx in (1, 2, 3):
            ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
        for cel in ws2[header_row2]:
            cel.font = Font(bold=True, color="1F2937")
            cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for item in extras.get("afastamentos", []) or []:
            ws2.append([
                item.get("tipo", ""),
                _data_iso_para_display(item.get("data_inicio", "")),
                _data_iso_para_display(item.get("data_fim", "")),
                item.get("observacoes", ""),
                _data_hora_para_display(item.get("criado_em", "")),
            ])
        for idx, row in enumerate(ws2.iter_rows(min_row=header_row2 + 1, max_row=ws2.max_row), start=1):
            for cel in row:
                cel.alignment = Alignment(vertical="top")
                cel.fill = PatternFill(start_color="F7FAFC" if idx % 2 else "EEF3F8", end_color="F7FAFC" if idx % 2 else "EEF3F8", fill_type="solid")
        ws2.freeze_panes = ws2["A{}".format(header_row2 + 1)]
        ws2.auto_filter.ref = "A{}:E{}".format(header_row2, ws2.max_row)
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 14
        ws2.column_dimensions["D"].width = 34
        ws2.column_dimensions["E"].width = 18

        ws3 = wb.create_sheet("Atestados")
        ws3.sheet_properties.tabColor = "FF6B35"
        ws3.append(["Atestados"])
        ws3.append(["Funcionario: {}".format(funcionario.get("name", ""))])
        ws3.append(["Matricula: {}".format(funcionario.get("enrollid", ""))])
        ws3.append([])
        ws3.append(["Arquivo", "Emissao", "Validade", "Observacoes", "Criado em"])
        header_row3 = ws3.max_row
        for row_idx in (1, 2, 3):
            ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
        for cel in ws3[header_row3]:
            cel.font = Font(bold=True, color="1F2937")
            cel.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for item in extras.get("atestados", []) or []:
            ws3.append([
                item.get("nome_original", ""),
                _data_iso_para_display(item.get("data_emissao", "")),
                _data_iso_para_display(item.get("validade_ate", "")),
                item.get("observacoes", ""),
                _data_hora_para_display(item.get("criado_em", "")),
            ])
        for idx, row in enumerate(ws3.iter_rows(min_row=header_row3 + 1, max_row=ws3.max_row), start=1):
            for cel in row:
                cel.alignment = Alignment(vertical="top")
                cel.fill = PatternFill(start_color="F7FAFC" if idx % 2 else "EEF3F8", end_color="F7FAFC" if idx % 2 else "EEF3F8", fill_type="solid")
        ws3.freeze_panes = ws3["A{}".format(header_row3 + 1)]
        ws3.auto_filter.ref = "A{}:E{}".format(header_row3, ws3.max_row)
        ws3.column_dimensions["A"].width = 32
        ws3.column_dimensions["B"].width = 14
        ws3.column_dimensions["C"].width = 14
        ws3.column_dimensions["D"].width = 34
        ws3.column_dimensions["E"].width = 18
        wb.save(caminho)
    else:
        titulo = "Perfil do Funcionario"
        subtitulo = "{} - Matricula {}".format(funcionario.get("name", ""), funcionario.get("enrollid", ""))
        linhas_pdf = []
        for campo, valor in resumo:
            linhas_pdf.append(["Cadastro", campo, str(valor), "", ""])
        for item in extras.get("afastamentos", []) or []:
            linhas_pdf.append([
                "Afastamento",
                item.get("tipo", ""),
                _data_iso_para_display(item.get("data_inicio", "")),
                _data_iso_para_display(item.get("data_fim", "")),
                item.get("observacoes", ""),
            ])
        for item in extras.get("atestados", []) or []:
            linhas_pdf.append([
                "Atestado",
                item.get("nome_original", ""),
                _data_iso_para_display(item.get("data_emissao", "")),
                _data_iso_para_display(item.get("validade_ate", "")),
                item.get("observacoes", ""),
            ])
        ponto_core._gerar_pdf_texto(
            caminho,
            titulo,
            subtitulo,
            ["Seccao", "Campo", "Inicio", "Fim", "Detalhe"],
            linhas_pdf,
        )

    return send_from_directory(output_dir, nome_arquivo, as_attachment=True)


@app.route("/funcionarios/<int:enrollid>/iniciar-face", methods=["POST"])
def funcionarios_iniciar_face(enrollid):
    config = config_atual()
    try:
        ponto_core.iniciar_cadastro_facial(config, enrollid)
        return jsonify({
            "ok": True,
            "mensagem": "Cadastro iniciado. Mantenha a pessoa de frente para a camera, sem mais de um rosto na imagem, de preferencia entre 50 e 80 cm do leitor.",
        })
    except DeviceError as e:
        return jsonify({"ok": False, "mensagem": str(e)})


@app.route("/funcionarios/status-face")
def funcionarios_status_face():
    config = config_atual()
    try:
        dados = ponto_core.consultar_status_cadastro_facial(config)
        status = dados.get("status") if dados else None
        msg = dados.get("msg") if dados else None
        return jsonify({"ok": True, "status": status, "mensagem": msg})
    except DeviceError as e:
        return jsonify({"ok": False, "mensagem": str(e)})


@app.route("/funcionarios/cancelar-face", methods=["POST"])
def funcionarios_cancelar_face():
    config = config_atual()
    try:
        ponto_core.cancelar_cadastro_facial(config)
        return jsonify({"ok": True})
    except DeviceError as e:
        return jsonify({"ok": False, "mensagem": str(e)})


@app.route("/funcionarios/<int:enrollid>/remover-face", methods=["POST"])
def funcionarios_remover_face(enrollid):
    config = config_atual()
    try:
        ponto_core.remover_rosto(config, enrollid)
        flash("Rosto cadastrado removido. A pessoa precisa cadastrar de novo pra continuar batendo ponto por reconhecimento facial.", "ok")
    except DeviceError as e:
        flash(str(e), "erro")
    return redirect(url_for("funcionarios"))


@app.route("/configuracoes", methods=["GET", "POST"])
def configuracoes():
    config = config_atual()

    if request.method == "POST":
        config["device_ip"] = request.form.get("device_ip", "").strip()
        try:
            config["device_port"] = int(request.form.get("device_port", "80"))
        except ValueError:
            config["device_port"] = 80
        nova_senha = request.form.get("device_password", "")
        if nova_senha:
            config["device_password"] = nova_senha
        config["company_name"] = request.form.get("company_name", "").strip() or "Minha Empresa"
        config["company_document"] = request.form.get("company_document", "").strip()
        config["workplace_address"] = request.form.get("workplace_address", "").strip()
        config["rep_identifier"] = request.form.get("rep_identifier", "").strip()
        config["output_dir"] = request.form.get("output_dir", "").strip() or config.get("output_dir", "relatorios")

        config["notificacoes_ativas"] = request.form.get("notificacoes_ativas") == "on"
        try:
            config["intervalo_notificacao_segundos"] = max(15, int(request.form.get("intervalo_notificacao_segundos", 60)))
        except ValueError:
            config["intervalo_notificacao_segundos"] = 60

        logo_file = request.files.get("logo")
        if logo_file and logo_file.filename:
            nome_seguro = secure_filename(logo_file.filename)
            logo_file.save(os.path.join(LOGOS_DIR, nome_seguro))
            config["logo_path"] = nome_seguro

        config_store.salvar(config)
        auditoria.registrar_evento("alteracao_configuracao", "Configuracoes gerais atualizadas")
        flash("Configuracoes salvas.", "ok")
        return redirect(url_for("configuracoes"))

    return render_template("configuracoes.html", config=config)


@app.route("/testar-conexao", methods=["POST"])
def testar_conexao_route():
    device_ip = request.form.get("device_ip", "").strip()
    try:
        device_port = int(request.form.get("device_port", "80"))
    except ValueError:
        device_port = 80
    device_password = request.form.get("device_password", "")

    config = config_atual()
    if not device_password:
        device_password = config.get("device_password", "")

    cfg = {"device_ip": device_ip, "device_port": device_port, "device_password": device_password}
    try:
        matriculas = ponto_core.get_user_ids(cfg)
        return jsonify({"ok": True, "mensagem": "Conectado! {} funcionario(s) encontrado(s) no leitor.".format(len(matriculas))})
    except DeviceError as e:
        return jsonify({"ok": False, "mensagem": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "mensagem": "Erro inesperado: {}".format(e)})


@app.route("/testar-notificacao", methods=["POST"])
def testar_notificacao():
    return jsonify({
        "ok": True,
        "mensagem": "As notificações do Windows foram canceladas. Use o dashboard ao vivo para acompanhar as batidas.",
    })


def _abrir_navegador():
    time.sleep(1.2)
    webbrowser.open("http://localhost:5000")


if __name__ == "__main__":
    threading.Thread(target=_abrir_navegador, daemon=True).start()
    # host 0.0.0.0 permite acessar de outros computadores na mesma rede,
    # se so voce for usar, pode trocar pra "127.0.0.1"
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
