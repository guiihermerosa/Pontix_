# -*- coding: utf-8 -*-

import csv
import json
import os
import sqlite3
import threading
import unicodedata
import requests
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

REQUEST_TIMEOUT_SECONDS = 15

COR_TITULO = "1F4E78"
COR_CABECALHO = "D9E1F2"
COR_FAIXA = "F7FAFC"
COR_FAIXA_ALT = "EEF3F8"
COR_BORDA = "B7C2D0"

FIN_THIN = Side(style="thin", color="B7B7B7")
BORDA_PADRAO = Border(left=FIN_THIN, right=FIN_THIN, top=FIN_THIN, bottom=FIN_THIN)
HISTORICO_PASTA = "historico_ponto"
HISTORICO_DB = "facial_logs.sqlite3"
HISTORICO_JSONL = "ponto_historico.jsonl"
HISTORICO_CSV = "ponto_historico.csv"
HISTORICO_ESTADO = "coletor_estado.json"
_HISTORICO_LOCK = threading.Lock()


class DeviceError(Exception):
    """Erro ao falar com o leitor facial. A mensagem ja vem pronta pra mostrar ao usuario."""
    pass


def _device_url(ip, port):
    return "http://{}:{}/api".format(ip, port)


def api_call(config, cmd, extra_fields=None):
    payload = {"password": config["device_password"], "cmd": cmd}
    if extra_fields:
        payload.update(extra_fields)

    url = _device_url(config["device_ip"], config["device_port"])

    try:
        resp = requests.post(url, json=payload, timeout=REQUEST_TIMEOUT_SECONDS)
    except requests.exceptions.ConnectTimeout:
        raise DeviceError(
            "Nao consegui conectar no leitor ({}). Ele esta ligado e na mesma rede?".format(url)
        )
    except requests.exceptions.ConnectionError:
        raise DeviceError(
            "Nao consegui conectar no leitor ({}). Confira o IP configurado.".format(url)
        )
    except requests.exceptions.RequestException as e:
        raise DeviceError("Falha na comunicacao com o leitor: {}".format(e))

    if resp.status_code != 200:
        raise DeviceError("O leitor respondeu com status HTTP {}.".format(resp.status_code))

    try:
        data = resp.json()
    except ValueError:
        raise DeviceError("O leitor nao devolveu uma resposta valida (JSON).")

    if isinstance(data, dict) and data.get("result") is False:
        raise DeviceError("O leitor recusou o pedido: {}".format(data.get("msg", "sem detalhes")))

    return data


def _resolver_saida(path_saida):
    if os.path.isabs(path_saida):
        return path_saida
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), path_saida)


def _historico_dir(output_dir):
    return os.path.join(_resolver_saida(output_dir), HISTORICO_PASTA)


def _garantir_diretorio(path):
    if path and not os.path.isdir(path):
        os.makedirs(path)


def _historico_jsonl_path(output_dir):
    return os.path.join(_historico_dir(output_dir), HISTORICO_JSONL)


def _historico_csv_path(output_dir):
    return os.path.join(_historico_dir(output_dir), HISTORICO_CSV)


def _historico_db_path(output_dir):
    return os.path.join(_historico_dir(output_dir), HISTORICO_DB)


def _historico_estado_path(output_dir):
    return os.path.join(_historico_dir(output_dir), HISTORICO_ESTADO)


def _assinatura_log_bruto(item):
    return "|".join([
        str(item.get("time", "")).strip(),
        str(item.get("data", "")).strip(),
        str(item.get("hora", "")).strip(),
        str(item.get("enrollid", "")).strip(),
        str(item.get("name", "")).strip(),
        str(item.get("mode", "")).strip(),
        str(item.get("inout", "")).strip(),
        str(item.get("event", "")).strip(),
        str(item.get("note", "")).strip(),
        str(item.get("photourl", "")).strip(),
    ])


def _limpar_texto_data_hora(texto):
    texto = str(texto or "").strip()
    if not texto:
        return ""
    partes = texto.split(" ", 1)
    if len(partes) == 2:
        data_txt, hora_txt = partes
        hora_txt = hora_txt.replace(" ", "")
        return "{} {}".format(data_txt, hora_txt)
    return texto.replace(" ", "")


def _normalizar_log_bruto(registro, source):
    item = dict(registro or {})
    item["source"] = source or item.get("source", "")
    item.setdefault("salvo_em", datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))

    time_txt = _limpar_texto_data_hora(item.get("time", ""))
    data_txt = str(item.get("data", "")).strip()
    hora_txt = str(item.get("hora", "")).strip()

    if time_txt and (not data_txt or not hora_txt):
        data_hora = _data_hora_from_texto(time_txt)
        if data_hora is not None:
            data_txt = data_txt or data_hora.strftime("%Y-%m-%d")
            hora_txt = hora_txt or data_hora.strftime("%H:%M:%S")
    elif data_txt and hora_txt and not time_txt:
        time_txt = "{} {}".format(data_txt, hora_txt)

    item["time"] = time_txt
    item["data"] = data_txt
    item["hora"] = hora_txt
    item["signature"] = _assinatura_log_bruto(item)
    return item


def _abrir_historico_db(output_dir):
    caminho = _historico_db_path(output_dir)
    _garantir_diretorio(os.path.dirname(caminho))
    db = sqlite3.connect(caminho)
    db.row_factory = sqlite3.Row
    return db


def _garantir_historico_db(output_dir):
    with _abrir_historico_db(output_dir) as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS facial_logs (
                signature TEXT PRIMARY KEY,
                salvo_em TEXT,
                source TEXT,
                time TEXT,
                data TEXT,
                hora TEXT,
                enrollid TEXT,
                name TEXT,
                department TEXT,
                shiftid TEXT,
                shift_name TEXT,
                mode TEXT,
                inout TEXT,
                event TEXT,
                note TEXT,
                photourl TEXT,
                payload_json TEXT
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS facial_adjustments (
                original_signature TEXT PRIMARY KEY,
                adjusted_acao TEXT,
                observacao TEXT,
                ajustado_em TEXT,
                ajustado_por TEXT,
                original_data TEXT,
                original_hora TEXT,
                enrollid TEXT,
                name TEXT,
                department TEXT,
                source TEXT
            )
            """
        )
        db.execute("CREATE INDEX IF NOT EXISTS idx_facial_logs_data_hora ON facial_logs(data, hora)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_facial_adjustments_data ON facial_adjustments(original_data, original_hora)")
        db.commit()


def _gravar_csv_historico(csv_path, item):
    campos_csv = [
        "salvo_em",
        "data",
        "hora",
        "enrollid",
        "name",
        "department",
        "source",
    ]

    novo_arquivo_csv = not os.path.isfile(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8-sig") as f_csv:
        writer = csv.DictWriter(f_csv, fieldnames=campos_csv)
        if novo_arquivo_csv:
            writer.writeheader()
        writer.writerow({campo: item.get(campo, "") for campo in campos_csv})


def _gravar_jsonl_historico(jsonl_path, item):
    with open(jsonl_path, "a", encoding="utf-8") as f_json:
        f_json.write(json.dumps(item, ensure_ascii=False) + "\n")


def ler_estado_coletor(config):
    output_dir = _resolver_saida(config["output_dir"])
    caminho = _historico_estado_path(output_dir)
    if not os.path.isfile(caminho):
        return {}
    try:
        with open(caminho, "r", encoding="utf-8") as f_estado:
            return json.load(f_estado) or {}
    except Exception:
        return {}


def salvar_estado_coletor(config, estado):
    output_dir = _resolver_saida(config["output_dir"])
    _garantir_diretorio(_historico_dir(output_dir))
    caminho = _historico_estado_path(output_dir)
    with open(caminho, "w", encoding="utf-8") as f_estado:
        json.dump(estado or {}, f_estado, ensure_ascii=False, indent=2)


def registrar_ajuste_local(config, original_signature, adjusted_acao, observacao, ajustado_por="web"):
    original_signature = str(original_signature or "").strip()
    adjusted_acao = _acao_normalizada(adjusted_acao)
    observacao = str(observacao or "").strip()
    ajustado_por = str(ajustado_por or "").strip() or "web"

    if not original_signature:
        raise ValueError("Nao encontrei o registro original para ajustar.")
    if adjusted_acao not in ("Entrada", "Saida"):
        raise ValueError("A acao ajustada precisa ser Entrada ou Saida.")
    if not observacao:
        raise ValueError("A observacao do ajuste e obrigatoria.")

    output_dir = _resolver_saida(config["output_dir"])
    _garantir_diretorio(_historico_dir(output_dir))
    _garantir_historico_db(output_dir)

    registro_base = {}
    historico = ler_historico_local(config, limite=None)
    for item in historico:
        if str(item.get("signature", "")).strip() == original_signature:
            registro_base = dict(item)
            break

    with _HISTORICO_LOCK:
        with _abrir_historico_db(output_dir) as db:
            db.execute(
                """
                INSERT OR REPLACE INTO facial_adjustments (
                    original_signature, adjusted_acao, observacao, ajustado_em, ajustado_por,
                    original_data, original_hora, enrollid, name, department, source
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    original_signature,
                    adjusted_acao,
                    observacao,
                    datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                    ajustado_por,
                    str(registro_base.get("data", "")),
                    str(registro_base.get("hora", "")),
                    str(registro_base.get("enrollid", "")),
                    str(registro_base.get("name", "")),
                    str(registro_base.get("department", "")),
                    str(registro_base.get("source", "")),
                ),
            )
            db.commit()

    return {
        "original_signature": original_signature,
        "acao": adjusted_acao,
        "observacao": observacao,
        "ajustado_por": ajustado_por,
    }


def listar_ajustes_locais(config, inicio=None, fim=None):
    output_dir = _resolver_saida(config["output_dir"])
    db_path = _historico_db_path(output_dir)
    if not os.path.isfile(db_path):
        return []

    # Garante migração automática do banco antigo antes de consultar ajustes.
    _garantir_historico_db(output_dir)

    inicio_txt = inicio.isoformat() if inicio else None
    fim_txt = fim.isoformat() if fim else None
    ajustes = []

    with sqlite3.connect(db_path) as db:
        db.row_factory = sqlite3.Row
        query = "SELECT * FROM facial_adjustments WHERE 1=1"
        params = []
        if inicio_txt:
            query += " AND original_data >= ?"
            params.append(inicio_txt)
        if fim_txt:
            query += " AND original_data <= ?"
            params.append(fim_txt)
        query += " ORDER BY original_data, original_hora, enrollid"
        for row in db.execute(query, params):
            ajustes.append(dict(row))
    return ajustes


def _mapa_ajustes_locais(config, inicio=None, fim=None):
    ajustes = {}
    for item in listar_ajustes_locais(config, inicio=inicio, fim=fim):
        ajustes[str(item.get("original_signature", "")).strip()] = item
    return ajustes


def _acao_normalizada(valor):
    valor = str(valor or "").strip().lower()
    if not valor:
        return ""
    if valor.startswith("entr"):
        return "Entrada"
    if valor.startswith("sai"):
        return "Saida"
    return ""


def _chave_acao_por_dia(data_txt, enrollid):
    return "{}|{}".format(str(data_txt or "").strip(), str(enrollid or "").strip())


def _ultima_acao_registrada(config, data_txt, enrollid):
    data_obj = None
    try:
        data_obj = datetime.strptime(str(data_txt or "").strip(), "%Y-%m-%d").date()
    except Exception:
        data_obj = None

    if not data_obj:
        return ""

    registros = ler_historico_local(config, inicio=data_obj, fim=data_obj)
    chave = _chave_acao_por_dia(data_obj.isoformat(), enrollid)
    for item in reversed(registros or []):
        if _chave_acao_por_dia(item.get("data", ""), item.get("enrollid", "")) != chave:
            continue
        acao = _acao_normalizada(item.get("acao"))
        if acao:
            return acao
    return ""


def determinar_acao_marcacao(config, registro, estado=None):
    """Alterna Entrada/Saida por funcionario e por dia."""
    item = dict(registro or {})
    data_txt = str(item.get("data", "")).strip()
    if not data_txt:
        data_txt = date.today().isoformat()

    matricula = str(item.get("enrollid", "")).strip()
    if not matricula:
        return "Entrada", estado or {}

    estado = dict(estado or {})
    mapa = estado.get("alternancia_marcacoes", {}) or {}
    if not isinstance(mapa, dict):
        mapa = {}

    chave = _chave_acao_por_dia(data_txt, matricula)
    ultima_acao = _acao_normalizada(mapa.get(chave))
    if not ultima_acao:
        ultima_acao = _ultima_acao_registrada(config, data_txt, matricula)

    if ultima_acao == "Entrada":
        nova_acao = "Saida"
    else:
        nova_acao = "Entrada"

    mapa[chave] = nova_acao
    estado["alternancia_marcacoes"] = mapa
    return nova_acao, estado


def log_local_ja_existe(config, assinatura):
    assinatura = str(assinatura or "").strip()
    if not assinatura:
        return False

    output_dir = _resolver_saida(config["output_dir"])
    db_path = _historico_db_path(output_dir)
    if os.path.isfile(db_path):
        with sqlite3.connect(db_path) as db:
            row = db.execute(
                "SELECT 1 FROM facial_logs WHERE signature = ? LIMIT 1",
                (assinatura,),
            ).fetchone()
            if row is not None:
                return True

    jsonl_path = _historico_jsonl_path(output_dir)
    if os.path.isfile(jsonl_path):
        with open(jsonl_path, "r", encoding="utf-8") as f_json:
            for linha in f_json:
                linha = linha.strip()
                if not linha:
                    continue
                try:
                    item = json.loads(linha)
                except ValueError:
                    continue
                if str(item.get("signature", "")).strip() == assinatura:
                    return True
    return False


def caminho_historico_csv(output_dir):
    return _historico_csv_path(output_dir)


def registrar_marcacao_local(config, registro):
    return registrar_log_local(config, registro, source=registro.get("source", "monitor"))


def registrar_log_local(config, registro, source="monitor"):
    """Salva qualquer log do facial em arquivo e banco locais."""
    output_dir = _resolver_saida(config["output_dir"])
    _garantir_diretorio(_historico_dir(output_dir))
    _garantir_historico_db(output_dir)
    item = _normalizar_log_bruto(registro, source)
    jsonl_path = _historico_jsonl_path(output_dir)
    csv_path = _historico_csv_path(output_dir)
    db_path = _historico_db_path(output_dir)

    with _HISTORICO_LOCK:
        with sqlite3.connect(db_path) as db:
            cur = db.execute(
                """
                INSERT OR IGNORE INTO facial_logs (
                    signature, salvo_em, source, time, data, hora, enrollid, name, department,
                    shiftid, shift_name, mode, inout, event, note, photourl, payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item.get("signature", ""),
                    item.get("salvo_em", ""),
                    item.get("source", ""),
                    item.get("time", ""),
                    item.get("data", ""),
                    item.get("hora", ""),
                    str(item.get("enrollid", "")),
                    item.get("name", ""),
                    item.get("department", ""),
                    str(item.get("shiftid", "")),
                    item.get("shift_name", ""),
                    str(item.get("mode", "")),
                    str(item.get("inout", "")),
                    str(item.get("event", "")),
                    item.get("note", ""),
                    item.get("photourl", ""),
                    json.dumps(item, ensure_ascii=False),
                ),
            )
            inserido = cur.rowcount > 0
            if not inserido:
                atual = db.execute(
                    "SELECT * FROM facial_logs WHERE signature = ?",
                    (item.get("signature", ""),),
                ).fetchone()
                if atual is not None:
                    merged = dict(atual)
                    payload_json_atual = merged.pop("payload_json", "")
                    if payload_json_atual:
                        try:
                            payload_atual = json.loads(payload_json_atual)
                        except ValueError:
                            payload_atual = {}
                        for chave, valor in payload_atual.items():
                            if not merged.get(chave):
                                merged[chave] = valor
                    for chave, valor in item.items():
                        if valor not in (None, ""):
                            merged[chave] = valor
                    db.execute(
                        """
                        UPDATE facial_logs SET
                            salvo_em = ?,
                            source = ?,
                            time = ?,
                            data = ?,
                            hora = ?,
                            enrollid = ?,
                            name = ?,
                            department = ?,
                            shiftid = ?,
                            shift_name = ?,
                            mode = ?,
                            inout = ?,
                            event = ?,
                            note = ?,
                            photourl = ?,
                            payload_json = ?
                        WHERE signature = ?
                        """,
                        (
                            merged.get("salvo_em", ""),
                            merged.get("source", ""),
                            merged.get("time", ""),
                            merged.get("data", ""),
                            merged.get("hora", ""),
                            str(merged.get("enrollid", "")),
                            merged.get("name", ""),
                            merged.get("department", ""),
                            str(merged.get("shiftid", "")),
                            merged.get("shift_name", ""),
                            str(merged.get("mode", "")),
                            str(merged.get("inout", "")),
                            str(merged.get("event", "")),
                            merged.get("note", ""),
                            merged.get("photourl", ""),
                            json.dumps(merged, ensure_ascii=False),
                            item.get("signature", ""),
                        ),
                    )
            db.commit()

        if inserido:
            _gravar_jsonl_historico(jsonl_path, item)
            _gravar_csv_historico(csv_path, item)
        item["gravado"] = inserido

    return item


def ler_historico_local(config, inicio=None, fim=None, limite=None):
    """Le o arquivo local de marcacoes salvas pelo monitoramento."""
    output_dir = _resolver_saida(config["output_dir"])
    jsonl_path = _historico_jsonl_path(output_dir)
    db_path = _historico_db_path(output_dir)

    inicio_txt = inicio.isoformat() if inicio else None
    fim_txt = fim.isoformat() if fim else None

    registros = []
    assinaturas = set()
    ajustes = _mapa_ajustes_locais(config, inicio=inicio, fim=fim)

    if os.path.isfile(db_path):
        with sqlite3.connect(db_path) as db:
            db.row_factory = sqlite3.Row
            query = "SELECT * FROM facial_logs WHERE 1=1"
            params = []
            if inicio_txt:
                query += " AND data >= ?"
                params.append(inicio_txt)
            if fim_txt:
                query += " AND data <= ?"
                params.append(fim_txt)
            query += " ORDER BY data, hora, enrollid"
            for row in db.execute(query, params):
                item = dict(row)
                payload_json = item.pop("payload_json", "")
                if payload_json:
                    try:
                        payload = json.loads(payload_json)
                    except ValueError:
                        payload = {}
                    for chave, valor in payload.items():
                        if item.get(chave) in (None, ""):
                            item[chave] = valor
                item.setdefault("data", "")
                item.setdefault("hora", "")
                item.setdefault("time", "")
                assinatura = item.get("signature") or _assinatura_log_bruto(item)
                if assinatura in assinaturas:
                    continue
                assinaturas.add(assinatura)
                ajuste = ajustes.get(str(assinatura).strip())
                if ajuste:
                    item["acao_original"] = _acao_normalizada(item.get("acao"))
                    item["acao"] = _acao_normalizada(ajuste.get("adjusted_acao")) or item.get("acao", "")
                    item["ajuste_manual"] = True
                    item["observacao_ajuste"] = ajuste.get("observacao", "")
                    item["ajustado_em"] = ajuste.get("ajustado_em", "")
                    item["ajustado_por"] = ajuste.get("ajustado_por", "")
                registros.append(item)

    if os.path.isfile(jsonl_path):
        with open(jsonl_path, "r", encoding="utf-8") as f_json:
            for linha in f_json:
                linha = linha.strip()
                if not linha:
                    continue
                try:
                    item = json.loads(linha)
                except ValueError:
                    continue

                data_item = str(item.get("data", ""))
                if inicio_txt and data_item < inicio_txt:
                    continue
                if fim_txt and data_item > fim_txt:
                    continue
                assinatura = item.get("signature") or _assinatura_log_bruto(item)
                if assinatura in assinaturas:
                    continue
                assinaturas.add(assinatura)
                ajuste = ajustes.get(str(assinatura).strip())
                if ajuste:
                    item["acao_original"] = _acao_normalizada(item.get("acao"))
                    item["acao"] = _acao_normalizada(ajuste.get("adjusted_acao")) or item.get("acao", "")
                    item["ajuste_manual"] = True
                    item["observacao_ajuste"] = ajuste.get("observacao", "")
                    item["ajustado_em"] = ajuste.get("ajustado_em", "")
                    item["ajustado_por"] = ajuste.get("ajustado_por", "")
                registros.append(item)

    registros.sort(key=lambda item: (
        str(item.get("data", "")),
        str(item.get("hora", "")),
        str(item.get("enrollid", "")),
    ))

    ultimas_acoes = {}
    for item in registros:
        data_item = str(item.get("data", "")).strip()
        matricula = str(item.get("enrollid", "")).strip()
        if not data_item or not matricula:
            continue
        chave = _chave_acao_por_dia(data_item, matricula)
        acao = _acao_normalizada(item.get("acao"))
        if not acao:
            if ultimas_acoes.get(chave) == "Entrada":
                acao = "Saida"
            else:
                acao = "Entrada"
            item["acao"] = acao
        ultimas_acoes[chave] = acao

    if limite is not None and limite > 0:
        registros = registros[-limite:]
    return registros


def gerar_fechamento_local(config, inicio, fim, nome_periodo):
    """Gera um XLSX baseado nas marcacoes locais salvas em arquivo."""
    registros = ler_historico_local(config, inicio=inicio, fim=fim)
    if not registros:
        raise DeviceError("Nao encontrei marcacoes salvas localmente para esse periodo.")

    turnos = turnos_com_config(config)
    regras = _regras_ponto(config)
    banco_cfg = regras.get("banco_horas", {}) or {}
    extras_cfg = regras.get("horas_extras", {}) or {}
    noturno_cfg = regras.get("adicional_noturno", {}) or {}
    toler_cfg = regras.get("tolerancias", {}) or {}
    intervalo_cfg = regras.get("intervalo", {}) or {}

    tolerancia_entrada = int(toler_cfg.get("entrada_minutos", 5) or 0)
    tolerancia_saida = int(toler_cfg.get("saida_minutos", 10) or 0)
    intervalo_minimo = int(intervalo_cfg.get("minimo_minutos", 30) or 0)
    intervalo_maximo = int(intervalo_cfg.get("maximo_minutos", 120) or 0)

    grupos = {}
    for item in registros:
        enrollid = str(item.get("enrollid", "")).strip()
        if not enrollid:
            continue
        grupo = grupos.setdefault(enrollid, {
            "enrollid": enrollid,
            "name": item.get("name", ""),
            "department": item.get("department", ""),
            "shiftid": int(item.get("shiftid", 0) or 0),
            "shift_name": item.get("shift_name", "") or _nome_turno(turnos, item.get("shiftid", 0)),
            "dias": {},
        })
        dia = str(item.get("data", "")).strip()
        hora = str(item.get("hora", "")).strip()
        if not dia:
            continue
        grupo["dias"].setdefault(dia, []).append(hora)

    if not grupos:
        raise DeviceError("O arquivo local nao tem dados validos para esse periodo.")

    resultados = []
    for grupo in sorted(grupos.values(), key=lambda g: (g.get("name", ""), g.get("enrollid", ""))):
        turno = {}
        shiftid = int(grupo.get("shiftid", 0) or 0)
        if 0 <= shiftid < len(turnos):
            turno = turnos[shiftid] or {}
        secoes_previstas = _turno_previsto(turno)
        horario_inicio_previsto = secoes_previstas[0][0] if secoes_previstas else ""
        horario_fim_previsto = secoes_previstas[-1][1] if secoes_previstas else ""
        minutos_previstos = 0
        for inicio_secao, fim_secao in secoes_previstas:
            inicio_m = _hora_para_minutos(inicio_secao)
            fim_m = _hora_para_minutos(fim_secao)
            if inicio_m is None or fim_m is None:
                continue
            minutos_previstos += max(0, fim_m - inicio_m)

        minutos_trabalhados = 0
        atraso_total = 0
        saida_antecipada_total = 0
        atraso_qtd = 0
        saida_antecipada_qtd = 0
        intervalo_real_total = 0
        intervalo_alertas = 0
        noturno_total = 0
        extra_50 = 0
        extra_70 = 0
        extra_100 = 0
        observacoes_periodo = []
        saldo_banco_total = 0

        dias = grupo["dias"]
        total_marcacoes = 0
        todas_horas = []
        dias_trabalhados = 0
        for dia_txt, lista_horas in dias.items():
            try:
                dia_obj = datetime.strptime(dia_txt, "%Y-%m-%d").date()
            except ValueError:
                continue
            marcacoes = []
            for hora in lista_horas:
                if _hora_para_minutos(hora) is not None:
                    marcacoes.append(hora)
            marcacoes.sort()
            if not marcacoes:
                continue
            dias_trabalhados += 1
            total_marcacoes += len(marcacoes)
            todas_horas.extend(marcacoes)
            minutos_trabalhados_dia = 0
            noturno_dia = 0

            primeira = marcacoes[0]
            ultima = marcacoes[-1]
            primeira_m = _hora_para_minutos(primeira)
            ultima_m = _hora_para_minutos(ultima)
            prevista_ini_m = _hora_para_minutos(horario_inicio_previsto)
            prevista_fim_m = _hora_para_minutos(horario_fim_previsto)

            if prevista_ini_m is not None and primeira_m is not None and primeira_m > prevista_ini_m + tolerancia_entrada:
                atraso_total += primeira_m - prevista_ini_m
                atraso_qtd += 1
            if prevista_fim_m is not None and ultima_m is not None and ultima_m < prevista_fim_m - tolerancia_saida:
                saida_antecipada_total += prevista_fim_m - ultima_m
                saida_antecipada_qtd += 1

            pares = zip(marcacoes[0::2], marcacoes[1::2])
            for entrada, saida in pares:
                entrada_m = _hora_para_minutos(entrada)
                saida_m = _hora_para_minutos(saida)
                if entrada_m is None or saida_m is None:
                    continue
                duracao = max(0, saida_m - entrada_m)
                minutos_trabalhados_dia += duracao
                if noturno_cfg.get("ativo", True):
                    noturno_dia += _intersecao_minutos(entrada, saida, noturno_cfg.get("inicio", "22:00"), noturno_cfg.get("fim", "05:00"))

            if len(marcacoes) >= 4:
                intervalo_real = _intervalo_minutos(marcacoes[1], marcacoes[2])
                intervalo_real_total += intervalo_real
                if intervalo_cfg.get("avisar", True) and (intervalo_real < intervalo_minimo or (intervalo_maximo > 0 and intervalo_real > intervalo_maximo)):
                    intervalo_alertas += 1
                    observacoes_periodo.append("{}: intervalo fora da regra ({} min)".format(dia_txt, intervalo_real))

            minutos_trabalhados += minutos_trabalhados_dia
            noturno_total += noturno_dia

            extra_dia = max(0, minutos_trabalhados_dia - minutos_previstos)
            classificacao = _calcular_classificacao_extra(config, dia_obj, extra_dia, noturno_dia)
            extra_50 += classificacao.get("50", 0)
            extra_70 += classificacao.get("70", 0)
            extra_100 += classificacao.get("100", 0)

            saldo_dia = minutos_trabalhados_dia - minutos_previstos
            if banco_cfg.get("ativo", True) and saldo_dia > 0:
                compensacao = int(banco_cfg.get("compensacao_minutos", 0) or 0)
                saldo_dia = max(0, saldo_dia - compensacao)
            saldo_banco_total += saldo_dia

        primeiras = min(todas_horas) if todas_horas else ""
        ultimas = max(todas_horas) if todas_horas else ""
        total_dias_periodo = (fim - inicio).days + 1
        dias_sem_ponto = max(0, total_dias_periodo - dias_trabalhados)
        if banco_cfg.get("limite_minutos") and saldo_banco_total > int(banco_cfg.get("limite_minutos", 0) or 0):
            observacoes_periodo.append("saldo acima do limite de banco de horas")
        if banco_cfg.get("validade_dias"):
            validade = int(banco_cfg.get("validade_dias", 0) or 0)
            if validade > 0 and (fim - inicio).days + 1 > validade:
                observacoes_periodo.append("periodo acima da validade configurada")
        resultados.append({
            "enrollid": grupo["enrollid"],
            "name": grupo["name"],
            "department": grupo["department"],
            "shift_name": grupo.get("shift_name", ""),
            "shiftid": shiftid,
            "dias_com_ponto": dias_trabalhados,
            "total_marcacoes": total_marcacoes,
            "primeira_marcacao": primeiras,
            "ultima_marcacao": ultimas,
            "work_hour": _minutos_para_hhmm(minutos_trabalhados),
            "atten_hour": _minutos_para_hhmm(minutos_previstos),
            "late_times": atraso_qtd,
            "late_minute": atraso_total,
            "leave_times": saida_antecipada_qtd,
            "leave_minute": saida_antecipada_total,
            "std_ot_hour": _minutos_para_hhmm(minutos_previstos),
            "ot_hour": _minutos_para_hhmm(max(0, minutos_trabalhados - minutos_previstos)),
            "bank_balance": _minutos_para_hhmm(saldo_banco_total),
            "bank_positive": _minutos_para_hhmm(max(0, saldo_banco_total)),
            "bank_negative": _minutos_para_hhmm(max(0, -saldo_banco_total)),
            "extra_50": _minutos_para_hhmm(extra_50),
            "extra_70": _minutos_para_hhmm(extra_70),
            "extra_100": _minutos_para_hhmm(extra_100),
            "night_minutes": _minutos_para_hhmm(noturno_total),
            "intervalo_total": _minutos_para_hhmm(intervalo_real_total),
            "intervalo_alertas": intervalo_alertas,
            "work_days": dias_trabalhados,
            "absent_days": dias_sem_ponto,
            "observacoes": _formatar_observacoes(observacoes_periodo),
            "dias": dias,
        })

    output_dir = _resolver_saida(config["output_dir"])
    _garantir_diretorio(output_dir)
    periodo_txt = "{} a {}".format(inicio.strftime("%d/%m/%Y"), fim.strftime("%d/%m/%Y"))
    logo_path = config.get("logo_path") or ""
    if logo_path and not os.path.isabs(logo_path):
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logos", logo_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Local"

    colunas = [
        "Matricula",
        "Nome",
        "Departamento",
        "Turno",
        "Dias com ponto",
        "Total marcacoes",
        "Primeira marcacao",
        "Ultima marcacao",
        "Horas previstas",
        "Horas trabalhadas",
        "Banco de horas",
        "Saldo positivo",
        "Saldo negativo",
        "Horas extras 50%",
        "Horas extras 70%",
        "Horas extras 100%",
        "Adicional noturno",
        "Intervalo total",
        "Alertas intervalo",
        "Atrasos (s)",
        "Saidas Antecip. (s)",
        "Faltas",
        "Observacoes",
    ]
    linha = escrever_cabecalho_pagina(ws, "Fechamento de Ponto - Resumo Local", periodo_txt, len(colunas), config["company_name"], logo_path)
    escrever_linha_cabecalho_tabela(ws, linha, colunas)
    linha += 1

    for item in resultados:
        valores = [
            item["enrollid"],
            item["name"],
            item["department"],
            item.get("shift_name", ""),
            item["dias_com_ponto"],
            item["total_marcacoes"],
            item["primeira_marcacao"],
            item["ultima_marcacao"],
            item["atten_hour"],
            item["work_hour"],
            item["bank_balance"],
            item["bank_positive"],
            item["bank_negative"],
            item["extra_50"],
            item["extra_70"],
            item["extra_100"],
            item["night_minutes"],
            item["intervalo_total"],
            item["intervalo_alertas"],
            _minutos_para_segundos(item["late_minute"]),
            _minutos_para_segundos(item["leave_minute"]),
            item["absent_days"],
            item["observacoes"],
        ]
        for col_i, valor in enumerate(valores, start=1):
            cel = ws.cell(row=linha, column=col_i, value=valor)
            cel.border = BORDA_PADRAO
            cel.alignment = Alignment(horizontal="center")
            if col_i in (21, 23):
                cel.value = _segundos_para_duracao_excel(cel.value)
                _formatar_duracao_excel(cel)
        linha += 1

    for col_i in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 18

    ws_mov = wb.create_sheet("Movimentos")
    colunas_mov = ["Data", "Hora", "Matricula", "Nome", "Departamento", "Turno", "Origem", "Salvo em"]
    linha_mov = escrever_cabecalho_pagina(ws_mov, "Fechamento de Ponto - Movimentos", periodo_txt, len(colunas_mov), config["company_name"], logo_path)
    escrever_linha_cabecalho_tabela(ws_mov, linha_mov, colunas_mov)
    linha_mov += 1

    for item in registros:
        valores = [
            item.get("data", ""),
            item.get("hora", ""),
            item.get("enrollid", ""),
            item.get("name", ""),
            item.get("department", ""),
            item.get("shift_name", ""),
            item.get("source", "monitor"),
            item.get("salvo_em", ""),
        ]
        for col_i, valor in enumerate(valores, start=1):
            cel = ws_mov.cell(row=linha_mov, column=col_i, value=valor)
            cel.border = BORDA_PADRAO
            cel.alignment = Alignment(horizontal="center")
        linha_mov += 1

    for col_i in range(1, len(colunas_mov) + 1):
        ws_mov.column_dimensions[get_column_letter(col_i)].width = 18

    nome_arquivo = "Fechamento_Ponto_{}_{}_a_{}.xlsx".format(
        nome_periodo,
        inicio.strftime("%Y-%m-%d"),
        fim.strftime("%Y-%m-%d"),
    )
    caminho_completo = os.path.join(output_dir, nome_arquivo)
    wb.save(caminho_completo)

    base_sem_extensao = os.path.splitext(caminho_completo)[0]
    cabecalhos_csv = colunas
    linhas_csv = []
    for item in resultados:
        linhas_csv.append([
            item["enrollid"],
            item["name"],
            item["department"],
            item.get("shift_name", ""),
            item["dias_com_ponto"],
            item["total_marcacoes"],
            item["primeira_marcacao"],
            item["ultima_marcacao"],
            item["atten_hour"],
            item["work_hour"],
            item["bank_balance"],
            item["bank_positive"],
            item["bank_negative"],
            item["extra_50"],
            item["extra_70"],
            item["extra_100"],
            item["night_minutes"],
            item["intervalo_total"],
            item["intervalo_alertas"],
            _minutos_para_segundos(item["late_minute"]),
            _minutos_para_segundos(item["leave_minute"]),
            item["absent_days"],
            item["observacoes"],
        ])

    _salvar_csv(base_sem_extensao + ".csv", cabecalhos_csv, linhas_csv)
    _gerar_pdf_texto(
        base_sem_extensao + ".pdf",
        "Fechamento de Ponto - Resumo Local",
        periodo_txt,
        cabecalhos_csv,
        linhas_csv,
    )
    return caminho_completo, nome_arquivo, len(registros)


def get_user_ids(config):
    data = api_call(config, "getuserids")
    if not data:
        return []
    return data.get("record", [])


def get_report(config, enrollid, from_date, to_date):
    return api_call(config, "getreport", {
        "enrollid": int(enrollid),
        "from": from_date,
        "to": to_date,
    })


def get_rtlog(config, index=0):
    return api_call(config, "getrtlog", {"index": int(index)})


def get_log(config, index=0, inicio=None, fim=None):
    payload = {"index": int(index)}
    if inicio is not None:
        payload["from"] = inicio.strftime("%y%m%d")
    if fim is not None:
        payload["to"] = fim.strftime("%y%m%d")
    return api_call(config, "getlog", payload)


def listar_logs_paginados(config, cmd_fn, limite=200, inicio=None, fim=None):
    registros = []
    index = 0
    seguranca = 0

    while True:
        dados = cmd_fn(config, index=index, inicio=inicio, fim=fim)
        if not dados:
            break

        itens = dados.get("record", []) or []
        for item in itens:
            registros.append(dict(item))
            if limite is not None and limite > 0 and len(registros) >= limite:
                return registros[:limite]

        if not itens:
            break

        try:
            total = int(dados.get("count", 0) or 0)
        except (TypeError, ValueError):
            total = 0
        try:
            ate = int(dados.get("to", 0) or 0)
        except (TypeError, ValueError):
            ate = 0

        if total and total <= ate:
            break

        # O firmware costuma devolver "to" como o próximo índice de leitura.
        # Somar +1 aqui faz pular um registro em alguns modelos, então usamos
        # o valor retornado pelo leitor como base e só caímos para o tamanho do lote
        # se o campo vier vazio ou inconsistente.
        proximo = ate if ate > index else index + len(itens)
        if proximo <= index:
            break

        index = proximo
        seguranca += 1
        if seguranca >= 50:
            break

    return registros


def listar_rtlogs(config, limite=200):
    """Lê os logs em tempo real do leitor facial usando paginação."""
    return listar_logs_paginados(config, lambda cfg, index=0, inicio=None, fim=None: get_rtlog(cfg, index=index), limite=limite)


def listar_getlogs(config, inicio, fim, limite=200):
    """Lê o histórico do leitor facial usando getlog."""
    return listar_logs_paginados(config, get_log, limite=limite, inicio=inicio, fim=fim)


def _data_hora_from_texto(texto):
    try:
        texto = _limpar_texto_data_hora(texto)
        return datetime.strptime(texto, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _dedupe_sequencia(valores):
    vistos = set()
    resultado = []
    for valor in valores:
        if valor in vistos:
            continue
        vistos.add(valor)
        resultado.append(valor)
    return resultado


def _normalizar_logs_para_relatorio(config, inicio, fim, logs, progress_cb=None):
    inicio_dt = datetime.combine(inicio, datetime.min.time())
    fim_dt = datetime.combine(fim, datetime.max.time())

    por_funcionario = {}
    for log in logs:
        data_hora = _data_hora_from_texto(log.get("time", ""))
        if data_hora is None or data_hora < inicio_dt or data_hora > fim_dt:
            continue

        enrollid = str(log.get("enrollid", "")).strip()
        if not enrollid:
            continue

        grupo = por_funcionario.setdefault(enrollid, {
            "enrollid": enrollid,
            "name": log.get("name", "") or "",
            "department": "",
            "shiftid": 0,
            "shift_name": "",
            "dias": {},
        })
        grupo["name"] = grupo.get("name") or (log.get("name", "") or "")
        dia_txt = data_hora.strftime("%Y-%m-%d")
        hora_txt = data_hora.strftime("%H:%M:%S")
        grupo["dias"].setdefault(dia_txt, []).append(hora_txt)

    if not por_funcionario:
        return []

    turnos = turnos_com_config(config)
    resultados = []
    ids = list(sorted(por_funcionario.keys(), key=lambda x: int(x) if str(x).isdigit() else str(x)))
    total = len(ids)
    semana_map = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SAB", 6: "DOM"}

    for indice, enrollid in enumerate(ids, start=1):
        grupo = por_funcionario[enrollid]
        info_usuario = get_user_info(config, enrollid) or {}
        shiftid = int(info_usuario.get("shiftid", 0) or 0)
        grupo["department"] = info_usuario.get("department", "") or ""
        grupo["shiftid"] = shiftid
        grupo["shift_name"] = info_usuario.get("shift_name", "") or _nome_turno(turnos, shiftid)

        turno = {}
        if 0 <= shiftid < len(turnos):
            turno = turnos[shiftid] or {}
        secoes_previstas = _turno_previsto(turno)
        horario_inicio_previsto = secoes_previstas[0][0] if secoes_previstas else ""
        horario_fim_previsto = secoes_previstas[-1][1] if secoes_previstas else ""
        minutos_previstos = 0
        for inicio_secao, fim_secao in secoes_previstas:
            inicio_m = _hora_para_minutos(inicio_secao)
            fim_m = _hora_para_minutos(fim_secao)
            if inicio_m is None or fim_m is None:
                continue
            minutos_previstos += max(0, fim_m - inicio_m)

        dias_trabalhados = 0
        total_marcacoes = 0
        minutos_trabalhados = 0
        atraso_total = 0
        saida_antecipada_total = 0
        atraso_qtd = 0
        saida_antecipada_qtd = 0
        intervalo_real_total = 0
        intervalo_alertas = 0
        noturno_total = 0
        extra_50 = 0
        extra_70 = 0
        extra_100 = 0
        observacoes_periodo = []
        saldo_banco_total = 0
        todas_horas = []
        dias_formatados = {}

        dias_itens = grupo["dias"]
        for dia_txt in sorted(dias_itens.keys()):
            horas = _dedupe_sequencia(sorted(dias_itens[dia_txt]))
            if not horas:
                continue
            dias_formatados[dia_txt] = horas
            dias_trabalhados += 1
            total_marcacoes += len(horas)
            todas_horas.extend(horas)

            dia_obj = datetime.strptime(dia_txt, "%Y-%m-%d").date()
            primeira = horas[0]
            ultima = horas[-1]
            primeira_m = _hora_para_minutos(primeira)
            ultima_m = _hora_para_minutos(ultima)
            prevista_ini_m = _hora_para_minutos(horario_inicio_previsto)
            prevista_fim_m = _hora_para_minutos(horario_fim_previsto)
            minutos_trabalhados_dia = 0
            noturno_dia = 0
            atraso_dia = 0
            saida_antecipada_dia = 0
            horas_previstas_dia = minutos_previstos

            if prevista_ini_m is not None and primeira_m is not None and primeira_m > prevista_ini_m + int(config.get("regras_ponto", {}).get("tolerancias", {}).get("entrada_minutos", 5) or 0):
                atraso_dia = primeira_m - prevista_ini_m
                atraso_total += atraso_dia
                atraso_qtd += 1
            if prevista_fim_m is not None and ultima_m is not None and ultima_m < prevista_fim_m - int(config.get("regras_ponto", {}).get("tolerancias", {}).get("saida_minutos", 10) or 0):
                saida_antecipada_dia = prevista_fim_m - ultima_m
                saida_antecipada_total += saida_antecipada_dia
                saida_antecipada_qtd += 1

            for entrada, saida in zip(horas[0::2], horas[1::2]):
                entrada_m = _hora_para_minutos(entrada)
                saida_m = _hora_para_minutos(saida)
                if entrada_m is None or saida_m is None:
                    continue
                duracao = max(0, saida_m - entrada_m)
                minutos_trabalhados_dia += duracao
                if config.get("regras_ponto", {}).get("adicional_noturno", {}).get("ativo", True):
                    noturno_dia += _intersecao_minutos(
                        entrada,
                        saida,
                        config.get("regras_ponto", {}).get("adicional_noturno", {}).get("inicio", "22:00"),
                        config.get("regras_ponto", {}).get("adicional_noturno", {}).get("fim", "05:00"),
                    )

            if len(horas) >= 4:
                intervalo_real = _intervalo_minutos(horas[1], horas[2])
                intervalo_real_total += intervalo_real
                intervalo_cfg = config.get("regras_ponto", {}).get("intervalo", {}) or {}
                intervalo_minimo = int(intervalo_cfg.get("minimo_minutos", 30) or 0)
                intervalo_maximo = int(intervalo_cfg.get("maximo_minutos", 120) or 0)
                if intervalo_cfg.get("avisar", True) and (intervalo_real < intervalo_minimo or (intervalo_maximo > 0 and intervalo_real > intervalo_maximo)):
                    intervalo_alertas += 1
                    observacoes_periodo.append("{}: intervalo fora da regra ({} min)".format(dia_txt, intervalo_real))

            minutos_trabalhados += minutos_trabalhados_dia
            noturno_total += noturno_dia

            extra_dia = max(0, minutos_trabalhados_dia - minutos_previstos)
            classificacao = _calcular_classificacao_extra(config, dia_obj, extra_dia, noturno_dia)
            extra_50 += classificacao.get("50", 0)
            extra_70 += classificacao.get("70", 0)
            extra_100 += classificacao.get("100", 0)

            banco_cfg = config.get("regras_ponto", {}).get("banco_horas", {}) or {}
            saldo_dia = minutos_trabalhados_dia - minutos_previstos
            if banco_cfg.get("ativo", True) and saldo_dia > 0:
                compensacao = int(banco_cfg.get("compensacao_minutos", 0) or 0)
                saldo_dia = max(0, saldo_dia - compensacao)
            saldo_banco_total += saldo_dia
            dias_formatados[dia_txt] = {
                "day": dia_obj.strftime("%m-%d"),
                "week": semana_map[dia_obj.weekday()],
                "time": horas,
                "fatten": _minutos_para_hhmm(horas_previstas_dia),
                "fact_atten": _minutos_para_hhmm(minutos_trabalhados_dia),
                "dayot": _minutos_para_hhmm(max(0, minutos_trabalhados_dia - horas_previstas_dia)),
                "daylate": str(atraso_dia),
                "dayLeaveearly": str(saida_antecipada_dia),
            }

        total_dias_periodo = (fim - inicio).days + 1
        dias_sem_ponto = max(0, total_dias_periodo - dias_trabalhados)
        banco_cfg = config.get("regras_ponto", {}).get("banco_horas", {}) or {}
        if banco_cfg.get("limite_minutos") and saldo_banco_total > int(banco_cfg.get("limite_minutos", 0) or 0):
            observacoes_periodo.append("saldo acima do limite de banco de horas")
        if banco_cfg.get("validade_dias"):
            validade = int(banco_cfg.get("validade_dias", 0) or 0)
            if validade > 0 and total_dias_periodo > validade:
                observacoes_periodo.append("periodo acima da validade configurada")

        resultados.append({
            "enrollid": grupo["enrollid"],
            "name": grupo["name"],
            "department": grupo["department"],
            "shift_name": grupo.get("shift_name", ""),
            "shiftid": shiftid,
            "dias_com_ponto": dias_trabalhados,
            "total_marcacoes": total_marcacoes,
            "primeira_marcacao": min(todas_horas) if todas_horas else "",
            "ultima_marcacao": max(todas_horas) if todas_horas else "",
            "work_hour": _minutos_para_hhmm(minutos_trabalhados),
            "atten_hour": _minutos_para_hhmm(minutos_previstos),
            "late_times": atraso_qtd,
            "late_minute": atraso_total,
            "leave_times": saida_antecipada_qtd,
            "leave_minute": saida_antecipada_total,
            "std_ot_hour": _minutos_para_hhmm(minutos_previstos),
            "ot_hour": _minutos_para_hhmm(max(0, minutos_trabalhados - minutos_previstos)),
            "bank_balance": _minutos_para_hhmm(saldo_banco_total),
            "bank_positive": _minutos_para_hhmm(max(0, saldo_banco_total)),
            "bank_negative": _minutos_para_hhmm(max(0, -saldo_banco_total)),
            "extra_50": _minutos_para_hhmm(extra_50),
            "extra_70": _minutos_para_hhmm(extra_70),
            "extra_100": _minutos_para_hhmm(extra_100),
            "night_minutes": _minutos_para_hhmm(noturno_total),
            "intervalo_total": _minutos_para_hhmm(intervalo_real_total),
            "intervalo_alertas": intervalo_alertas,
            "work_days": dias_trabalhados,
            "absent_days": dias_sem_ponto,
            "observacoes": _formatar_observacoes(observacoes_periodo),
            "days": list(dias_formatados.values()),
        })

        if progress_cb:
            progress_cb(indice, total)

    return resultados


def _construir_resultados_de_getlogs(config, inicio, fim, progress_cb=None):
    logs = listar_getlogs(config, inicio, fim, limite=5000)
    if not logs:
        return []
    return _normalizar_logs_para_relatorio(config, inicio, fim, logs, progress_cb=progress_cb)


def _construir_resultados_de_historico_local(config, inicio, fim, progress_cb=None):
    registros = ler_historico_local(config, inicio=inicio, fim=fim, limite=5000)
    if not registros:
        return []

    logs = []
    for item in registros:
        data = str(item.get("data", "")).strip()
        hora = str(item.get("hora", "")).strip()
        if not data or not hora:
            continue
        logs.append({
            "enrollid": item.get("enrollid", ""),
            "name": item.get("name", ""),
            "time": "{} {}".format(data, hora),
            "mode": 8,
            "inout": item.get("inout", 0),
            "event": item.get("event", 0),
        })

    if not logs:
        return []

    return _normalizar_logs_para_relatorio(config, inicio, fim, logs, progress_cb=progress_cb)


def get_shifts(config):
    data = api_call(config, "getshift")
    if not data:
        return []
    return data.get("shifts", []) or []


def set_shifts(config, shifts):
    return api_call(config, "setshift", {"shifts": shifts})


def _default_holidays():
    return [{
        "name": "New Year",
        "startday": "01-01",
        "endday": "01-01",
        "shift": 0,
        "dayzone": 0,
    }]


def _default_belltimes():
    return []


def _default_devlock():
    return {
        "dayzone": [{"day": [{"section": "00:00~00:00"} for _ in range(5)]} for _ in range(8)],
        "weekzone": [{"week": [{"day": 0} for _ in range(7)]} for _ in range(8)],
        "nopentime": [{"day": 0} for _ in range(7)],
        "visitortime": [{"day": 0} for _ in range(7)],
    }


def get_holidays(config):
    data = api_call(config, "getholiday")
    if not data:
        return []
    return data.get("holidays", []) or []


def set_holidays(config, holidays):
    return api_call(config, "setholiday", {"holidays": _default_holidays() if holidays is None else holidays})


def get_belltimes(config):
    data = api_call(config, "getbelltime")
    if not data:
        return []
    return data.get("belltimes", []) or []


def set_belltimes(config, belltimes):
    return api_call(config, "setbelltime", {"belltimes": _default_belltimes() if belltimes is None else belltimes})


def get_devlock(config):
    data = api_call(config, "getdevlock")
    if not data:
        return {}
    return data


def set_devlock(config, devlock):
    payload = _default_devlock() if devlock is None else devlock
    return api_call(config, "setdevlock", {
        "dayzone": payload.get("dayzone", _default_devlock()["dayzone"]),
        "weekzone": payload.get("weekzone", _default_devlock()["weekzone"]),
        "nopentime": payload.get("nopentime", _default_devlock()["nopentime"]),
        "visitortime": payload.get("visitortime", _default_devlock()["visitortime"]),
    })


def _nome_turno(turnos, shiftid):
    try:
        indice = int(shiftid)
    except (TypeError, ValueError):
        indice = 0
    if 0 <= indice < len(turnos):
        return turnos[indice].get("name", "shift{}".format(indice + 1))
    return "shift{}".format(indice + 1)


def _hora_para_minutos(hora):
    try:
        partes = str(hora).split(":")
        return int(partes[0]) * 60 + int(partes[1])
    except Exception:
        return None


def _minutos_para_hhmm(total_minutos):
    if total_minutos is None:
        return ""
    total_minutos = int(round(total_minutos))
    horas, minutos = divmod(max(0, total_minutos), 60)
    return "{:02d}:{:02d}".format(horas, minutos)


def _minutos_para_segundos(total_minutos):
    if total_minutos is None:
        return ""
    return int(round(float(total_minutos) * 60))


def _segundos_para_duracao_excel(total_segundos):
    if total_segundos in (None, ""):
        return ""
    try:
        total_segundos = int(round(float(total_segundos)))
    except (TypeError, ValueError):
        return ""
    return timedelta(seconds=max(0, total_segundos))


def _formatar_duracao_excel(celula):
    celula.number_format = "[h]:mm:ss"


def turnos_com_config(config, turnos=None):
    """Combina os turnos vindos do leitor com as flags locais de ativacao."""
    turnos_base = list(turnos if turnos is not None else get_shifts(config))
    while len(turnos_base) < 8:
        turnos_base.append({
            "name": "shift{}".format(len(turnos_base) + 1),
            "cutoftime": "00:00",
            "sections": [],
        })
    flags = config.get("turnos_config", []) or []
    turnos_view = []

    for indice, turno in enumerate(turnos_base):
        turno_dict = dict(turno or {})
        flag_turno = flags[indice] if indice < len(flags) and isinstance(flags[indice], dict) else {}
        flags_secoes = flag_turno.get("sections", []) or []
        secoes = []

        for sec_index, secao in enumerate(turno_dict.get("sections", []) or []):
            secao_dict = dict(secao or {})
            inicio = str(secao_dict.get("start", "00:00")).strip() or "00:00"
            fim = str(secao_dict.get("end", "00:00")).strip() or "00:00"
            ativo_padrao = not (inicio == "00:00" and fim == "00:00")
            if sec_index < len(flags_secoes) and isinstance(flags_secoes[sec_index], dict):
                ativo = bool(flags_secoes[sec_index].get("active", ativo_padrao))
            else:
                ativo = bool(secao_dict.get("active", ativo_padrao))
            secao_dict["start"] = inicio
            secao_dict["end"] = fim
            secao_dict["type"] = int(secao_dict.get("type", 0) or 0)
            secao_dict["active"] = ativo
            secao_dict["label"] = "Principal" if sec_index == 0 else "Complementar"
            secoes.append(secao_dict)

        while len(secoes) < 3:
            secoes.append({
                "start": "00:00",
                "end": "00:00",
                "type": 0,
                "active": False,
                "label": "Complementar",
            })

        turno_dict["sections"] = secoes[:3]
        turno_dict["faixas_ativas"] = sum(1 for secao in turno_dict["sections"] if secao.get("active") and not (secao.get("start") == "00:00" and secao.get("end") == "00:00"))
        turno_dict["numero"] = indice + 1
        turnos_view.append(turno_dict)

    return turnos_view


def _turno_previsto(turno):
    seccoes = []
    for secao in turno.get("sections", []) or []:
        inicio = secao.get("start", "")
        fim = secao.get("end", "")
        ativo = secao.get("active")
        if ativo is None:
            ativo = not (str(inicio) == "00:00" and str(fim) == "00:00")
        if not ativo:
            continue
        if str(inicio) == "00:00" and str(fim) == "00:00":
            continue
        if not inicio or not fim:
            continue
        seccoes.append((inicio, fim))
    return seccoes


def _regras_ponto(config):
    return config.get("regras_ponto", {}) or {}


def _lista_feriados(config):
    regras = _regras_ponto(config)
    feriados = []
    for item in regras.get("feriados", []) or []:
        if item.get("ativo", True):
            feriados.append(item)
    return feriados


def _eh_feriado(config, dia):
    data_txt = dia.strftime("%Y-%m-%d")
    data_mmdd = dia.strftime("%m-%d")
    for feriado in _lista_feriados(config):
        valor = str(feriado.get("data", "")).strip()
        if not valor:
            continue
        if valor in (data_txt, data_mmdd):
            return True, feriado
    return False, None


def _intervalo_minutos(inicio_txt, fim_txt):
    inicio = _hora_para_minutos(inicio_txt)
    fim = _hora_para_minutos(fim_txt)
    if inicio is None or fim is None:
        return 0
    if fim >= inicio:
        return fim - inicio
    return (24 * 60 - inicio) + fim


def _intersecao_minutos(inicio1, fim1, inicio2, fim2):
    i1 = _hora_para_minutos(inicio1)
    f1 = _hora_para_minutos(fim1)
    i2 = _hora_para_minutos(inicio2)
    f2 = _hora_para_minutos(fim2)
    if None in (i1, f1, i2, f2):
        return 0

    intervalos1 = [(i1, f1)] if f1 >= i1 else [(i1, 24 * 60), (0, f1)]
    intervalos2 = [(i2, f2)] if f2 >= i2 else [(i2, 24 * 60), (0, f2)]
    total = 0
    for a_ini, a_fim in intervalos1:
        for b_ini, b_fim in intervalos2:
            ini = max(a_ini, b_ini)
            fim = min(a_fim, b_fim)
            if fim > ini:
                total += fim - ini
    return total


def _eh_domingo(dia):
    return dia.weekday() == 6


def _calcular_classificacao_extra(config, dia, extra_minutos, adicional_noturno_minutos):
    regras = _regras_ponto(config)
    horas_extras = regras.get("horas_extras", {}) or {}
    if not horas_extras.get("ativo", True):
        return {"50": 0, "70": 0, "100": 0}
    if extra_minutos <= 0:
        return {"50": 0, "70": 0, "100": 0}

    eh_feriado, _ = _eh_feriado(config, dia)
    if eh_feriado:
        return {"50": 0, "70": 0, "100": extra_minutos}
    if _eh_domingo(dia):
        return {"50": 0, "70": 0, "100": extra_minutos}
    if adicional_noturno_minutos > 0:
        return {"50": 0, "70": extra_minutos, "100": 0}
    return {"50": extra_minutos, "70": 0, "100": 0}


def _formatar_observacoes(observacoes):
    if not observacoes:
        return ""
    return "; ".join(observacoes)


def _normalizar_texto_busca(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip())
    texto = texto.encode("ascii", "ignore").decode("ascii")
    return texto.lower()


def _resolver_periodo_relatorio(inicio, fim, filtros=None):
    filtros = filtros or {}
    inicio_filtro = filtros.get("periodo_inicio")
    fim_filtro = filtros.get("periodo_fim")

    if inicio_filtro:
        try:
            inicio_filtro = datetime.strptime(str(inicio_filtro), "%Y-%m-%d").date()
        except ValueError:
            raise ValueError("Data inicial invalida. Use o formato YYYY-MM-DD.")
        if inicio is None or inicio_filtro > inicio:
            inicio = inicio_filtro

    if fim_filtro:
        try:
            fim_filtro = datetime.strptime(str(fim_filtro), "%Y-%m-%d").date()
        except ValueError:
            raise ValueError("Data final invalida. Use o formato YYYY-MM-DD.")
        if fim is None or fim_filtro < fim:
            fim = fim_filtro

    if inicio and fim and fim < inicio:
        raise ValueError("A data final precisa ser depois da data inicial.")
    return inicio, fim


def filtrar_resultados_relatorio(resultados, filtros=None):
    filtros = filtros or {}
    matricula = str(filtros.get("matricula", "") or "").strip()
    nome = _normalizar_texto_busca(filtros.get("nome", ""))
    departamento = _normalizar_texto_busca(filtros.get("departamento", ""))
    turno = _normalizar_texto_busca(filtros.get("turno", ""))
    status = _normalizar_texto_busca(filtros.get("status", ""))

    filtrados = []
    for item in resultados or []:
        matricula_item = str(item.get("enrollid", "") or "").strip()
        nome_item = _normalizar_texto_busca(item.get("name", ""))
        departamento_item = _normalizar_texto_busca(item.get("department", ""))
        turno_item = _normalizar_texto_busca(item.get("shift_name", ""))
        status_item = _normalizar_texto_busca(item.get("status", "") or item.get("situacao", ""))

        if matricula and matricula_item != matricula:
            continue
        if nome and nome not in nome_item:
            continue
        if departamento and departamento not in departamento_item:
            continue
        if turno and turno not in turno_item:
            continue
        if status and status not in status_item:
            continue
        filtrados.append(item)
    return filtrados


def testar_conexao(config):
    """Usado pela tela de configuracoes pra validar IP/senha antes de salvar."""
    get_user_ids(config)
    return True


# ---------------------------------------------------------------------
# Cadastro de funcionarios (matricula/nome/departamento no leitor)
# A biometria facial em si (a foto do rosto) so pode ser cadastrada no
# proprio equipamento, olhando pra camera. Aqui so criamos/editamos o
# "perfil" (matricula, nome, departamento) que o leitor usa nos relatorios.
# ---------------------------------------------------------------------

def get_next_enrollid(config):
    """Pergunta ao leitor qual a proxima matricula livre."""
    data = api_call(config, "getunuserdid")
    return data.get("enrollid") if data else None


def check_user_id(config, enrollid):
    """Verifica se uma matricula ja esta em uso no leitor."""
    data = api_call(config, "checkuserid", {"enrollid": int(enrollid)})
    if not data:
        return None
    return bool(data.get("exists"))


def _erro_usuario_nao_encontrado(erro):
    texto = str(erro).lower()
    return "can not find the user" in texto or "nao encontrou" in texto or "não encontrou" in texto


def _erro_sem_logs(erro):
    texto = str(erro).lower()
    return (
        "can not find logs" in texto
        or "can not find the log" in texto
        or "nao encontrou logs" in texto
        or "não encontrou logs" in texto
        or "sem registros" in texto
    )


def set_user_info(config, enrollid, name, department="", shiftid=0):
    """Cria (ou atualiza) o perfil de um funcionario no leitor: matricula,
    nome e departamento. Nao cadastra a face - isso e feito no equipamento."""
    return api_call(config, "setuserinfo", {
        "enrollid": int(enrollid),
        "name": name,
        "department": department or "",
        "shiftid": int(shiftid) if shiftid is not None else 0,
        "admin": 0,
        "pwd": 0,
        "card": 0,
        "zoneid": 0,
        "groupid": 0,
        "verifymode": 0,
    })


def get_user_info(config, enrollid):
    """Busca nome/departamento de uma matricula ja cadastrada.
    Devolve None se a matricula nao tiver um perfil completo."""
    try:
        data = api_call(config, "getuserinfo", {"enrollid": int(enrollid)})
    except DeviceError as e:
        if _erro_usuario_nao_encontrado(e):
            return None
        raise
    return data


def listar_funcionarios(config):
    """Monta a lista de funcionarios cadastrados (matricula + nome + depto),
    lendo a lista de matriculas e depois o perfil de cada uma."""
    matriculas = get_user_ids(config)
    turnos = turnos_com_config(config)
    funcionarios = []
    for enrollid in matriculas:
        info = get_user_info(config, enrollid)
        if info:
            shiftid = info.get("shiftid", 0)
            funcionarios.append({
                "enrollid": info.get("enrollid", enrollid),
                "name": info.get("name", ""),
                "department": info.get("department", ""),
                "shiftid": shiftid,
                "shift_name": _nome_turno(turnos, shiftid),
            })
        else:
            funcionarios.append({
                "enrollid": enrollid,
                "name": "(sem perfil)",
                "department": "",
                "shiftid": 0,
                "shift_name": _nome_turno(turnos, 0),
            })
    return funcionarios


# ---------------------------------------------------------------------
# Cadastro do rosto (biometria facial)
# O leitor NAO aceita enviar uma foto pronta - o cadastro precisa ser
# feito com a pessoa de frente pra camera do equipamento. O que da pra
# fazer remotamente e iniciar esse modo de cadastro e acompanhar o
# andamento, sem precisar ir ate a tela do proprio leitor.
# ---------------------------------------------------------------------

def iniciar_cadastro_facial(config, enrollid):
    """Coloca o leitor em modo de cadastro de rosto pra essa matricula.
    A pessoa precisa estar de frente pra camera do equipamento agora."""
    return api_call(config, "adduser", {"flag": 2, "enrollid": int(enrollid), "backupnum": 50})


def consultar_status_cadastro_facial(config):
    """Pergunta como esta o andamento do cadastro de rosto iniciado.
    status 0 = ainda capturando. Qualquer outro valor = terminou
    (sucesso ou falha/tempo esgotado)."""
    return api_call(config, "checkregstatus")


def cancelar_cadastro_facial(config):
    return api_call(config, "adduser", {"cancel": True})


def remover_rosto(config, enrollid):
    """Apaga o rosto cadastrado de uma matricula (a pessoa some do
    reconhecimento facial ate cadastrar de novo). Nao apaga o perfil
    (nome/departamento), so a biometria."""
    return api_call(config, "deleteuserface", {"enrollid": int(enrollid)})


# ---------------------------------------------------------------------
# Periodos
# ---------------------------------------------------------------------

def periodo_semana_passada():
    hoje = date.today()
    segunda_atual = hoje - timedelta(days=hoje.weekday())
    inicio = segunda_atual - timedelta(days=7)
    fim = segunda_atual - timedelta(days=1)
    return inicio, fim


def periodo_mes_passado():
    hoje = date.today()
    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_passado = primeiro_dia_mes_atual - timedelta(days=1)
    inicio = ultimo_dia_mes_passado.replace(day=1)
    return inicio, ultimo_dia_mes_passado


def periodo_semana_atual():
    hoje = date.today()
    inicio = hoje - timedelta(days=hoje.weekday())
    return inicio, hoje


def periodo_mes_atual():
    hoje = date.today()
    return hoje.replace(day=1), hoje


# ---------------------------------------------------------------------
# Montagem da planilha
# ---------------------------------------------------------------------

def escrever_cabecalho_pagina(ws, titulo_relatorio, periodo_txt, n_colunas, company_name, logo_path):
    linha_titulo = 1
    coluna_inicial = 1

    if logo_path and os.path.isfile(logo_path):
        try:
            img = XLImage(logo_path)
            max_altura = 60
            if img.height > max_altura:
                escala = max_altura / img.height
                img.height = max_altura
                img.width = int(img.width * escala)
            ws.add_image(img, "A1")
            coluna_inicial = 3
            ws.row_dimensions[1].height = 46
        except Exception:
            pass

    col_letra_ini = get_column_letter(coluna_inicial)
    col_letra_fim = get_column_letter(max(n_colunas, coluna_inicial + 2))

    ws.merge_cells("{}{}:{}{}".format(col_letra_ini, linha_titulo, col_letra_fim, linha_titulo))
    cel = ws["{}{}".format(col_letra_ini, linha_titulo)]
    cel.value = company_name
    cel.font = Font(size=14, bold=True, color=COR_TITULO)
    cel.alignment = Alignment(horizontal="left", vertical="center")

    linha_titulo += 1
    ws.merge_cells("{}{}:{}{}".format(col_letra_ini, linha_titulo, col_letra_fim, linha_titulo))
    cel = ws["{}{}".format(col_letra_ini, linha_titulo)]
    cel.value = "{} — {}".format(titulo_relatorio, periodo_txt)
    cel.font = Font(size=11, italic=True, color="555555")
    cel.alignment = Alignment(horizontal="left", vertical="center")

    return linha_titulo + 2


def escrever_linha_cabecalho_tabela(ws, linha, titulos):
    for i, titulo in enumerate(titulos, start=1):
        cel = ws.cell(row=linha, column=i, value=titulo)
        cel.font = Font(bold=True, color="1F1F1F")
        cel.fill = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA_PADRAO


def escrever_cabecalho_pagina(ws, titulo_relatorio, periodo_txt, n_colunas, company_name, logo_path, vendor_name="GRB Tecnologia"):
    linha_titulo = 1
    coluna_inicial = 1

    if logo_path and os.path.isfile(logo_path):
        try:
            img = XLImage(logo_path)
            max_altura = 60
            if img.height > max_altura:
                escala = max_altura / img.height
                img.height = max_altura
                img.width = int(img.width * escala)
            ws.add_image(img, "A1")
            coluna_inicial = 3
            ws.row_dimensions[1].height = 46
        except Exception:
            pass

    col_letra_ini = get_column_letter(coluna_inicial)
    col_letra_fim = get_column_letter(max(n_colunas, coluna_inicial + 2))

    ws.merge_cells("{}{}:{}{}".format(col_letra_ini, linha_titulo, col_letra_fim, linha_titulo))
    cel = ws["{}{}".format(col_letra_ini, linha_titulo)]
    cel.value = "{}{}".format(company_name, "  •  {}".format(vendor_name) if vendor_name else "")
    cel.font = Font(size=14, bold=True, color=COR_TITULO)
    cel.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha_titulo].height = 22

    linha_titulo += 1
    ws.merge_cells("{}{}:{}{}".format(col_letra_ini, linha_titulo, col_letra_fim, linha_titulo))
    cel = ws["{}{}".format(col_letra_ini, linha_titulo)]
    cel.value = "{} — {}".format(titulo_relatorio, periodo_txt)
    cel.font = Font(size=11, italic=True, color="4A5568")
    cel.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha_titulo].height = 20

    linha_titulo += 1
    ws.merge_cells("{}{}:{}{}".format(col_letra_ini, linha_titulo, col_letra_fim, linha_titulo))
    cel = ws["{}{}".format(col_letra_ini, linha_titulo)]
    cel.value = "Desenvolvido por {}".format(vendor_name) if vendor_name else ""
    cel.font = Font(size=9, bold=True, color="7A8699")
    cel.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha_titulo].height = 18

    return linha_titulo + 2


def _salvar_csv(caminho, cabecalhos, linhas):
    _garantir_diretorio(os.path.dirname(caminho))
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        escritor = csv.writer(f, delimiter=";")
        escritor.writerow(cabecalhos)
        for linha in linhas:
            escritor.writerow(linha)


def _pdf_escape(texto):
    texto = "" if texto is None else str(texto)
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_fit_text(texto, max_chars):
    texto = _pdf_escape(texto)
    if max_chars <= 0:
        return ""
    if len(texto) <= max_chars:
        return texto
    if max_chars <= 3:
        return texto[:max_chars]
    return texto[:max_chars - 3].rstrip() + "..."


def _gerar_pdf_texto(caminho, titulo, subtitulo, cabecalhos, linhas):
    _garantir_diretorio(os.path.dirname(caminho))
    largura_pagina = 842
    altura_pagina = 595
    margem_x = 28
    topo = 32
    rodape = 24
    altura_titulo = 62
    altura_cabecalho = 22
    altura_linha = 18
    largura_util = largura_pagina - (margem_x * 2)

    todas_linhas = [list(cabecalhos)] + [list(linha) for linha in linhas]
    if not linhas:
        todas_linhas = [list(cabecalhos), ["Nenhum registro encontrado no periodo."] + [""] * (len(cabecalhos) - 1)]

    maximos = []
    for col_i, cab in enumerate(cabecalhos):
        maior = len(_pdf_escape(cab))
        for linha in linhas:
            if col_i < len(linha):
                maior = max(maior, len(_pdf_escape(linha[col_i])))
        maximos.append(min(max(maior, 6), 24))

    total_maximos = sum(maximos) or 1
    larguras = [max(32, int(round(largura_util * (valor / total_maximos)))) for valor in maximos]
    ajuste = largura_util - sum(larguras)
    if larguras:
        larguras[-1] += ajuste

    linhas_por_pagina = max(1, int((altura_pagina - topo - altura_titulo - altura_cabecalho - rodape) / altura_linha))
    paginas = [linhas[i:i + linhas_por_pagina] for i in range(0, len(linhas), linhas_por_pagina)] or [[]]

    objetos = []
    pagina_objetos = []
    for indice, pagina in enumerate(paginas):
        num_conteudo = 5 + indice * 2
        num_pagina = 6 + indice * 2
        pagina_objetos.append(num_pagina)

        comandos = []
        y = altura_pagina - topo

        def add(cmd):
            comandos.append(cmd)

        def texto(x, y_pos, txt, tamanho=8, bold=False):
            fonte = "/F2" if bold else "/F1"
            add("BT")
            add("%s %d Tf" % (fonte, tamanho))
            add("1 0 0 1 %d %d Tm" % (x, y_pos))
            add("(%s) Tj" % _pdf_escape(txt))
            add("ET")

        def retangulo(x, y_pos, w, h, fill=None, stroke=None):
            add("q")
            if fill:
                add("%s rg" % fill)
                add("%d %d %d %d re f" % (x, y_pos, w, h))
            if stroke:
                add("%s RG" % stroke)
                add("1 w")
                add("%d %d %d %d re S" % (x, y_pos, w, h))
            add("Q")

        # Header band
        retangulo(margem_x, y - 42, largura_util, 42, fill="0.12 0.27 0.41")
        texto(margem_x + 12, y - 17, _pdf_fit_text(titulo, 120), tamanho=15, bold=True)
        texto(margem_x + 12, y - 31, _pdf_fit_text(subtitulo, 120), tamanho=9, bold=False)

        # Meta line
        texto(margem_x + largura_util - 150, y - 17, "GRB Tecnologia", tamanho=9, bold=True)
        texto(margem_x + largura_util - 150, y - 31, "Relatorio de Ponto", tamanho=8, bold=False)

        y -= altura_titulo
        texto(margem_x, y, "Resumo", tamanho=10, bold=True)
        texto(margem_x + 60, y, "Pagina %d de %d" % (indice + 1, len(paginas)), tamanho=8, bold=False)
        y -= 8

        # Table header
        x = margem_x
        header_y = y - altura_cabecalho
        retangulo(margem_x, header_y, largura_util, altura_cabecalho, fill="0.86 0.90 0.95")
        for col_i, cab in enumerate(cabecalhos):
            largura = larguras[col_i]
            if col_i > 0:
                add("q")
                add("0.78 0.82 0.88 RG")
                add("1 w")
                add("%d %d %d %d re S" % (x, header_y, largura, altura_cabecalho))
                add("Q")
            texto(x + 3, header_y + 7, _pdf_fit_text(cab, max(6, int(largura / 4.2))), tamanho=7, bold=True)
            x += largura

        # Rows
        y_row = header_y
        for row_index, linha in enumerate(pagina):
            y_row -= altura_linha
            fill = "0.97 0.98 0.99" if row_index % 2 == 0 else "1 1 1"
            retangulo(margem_x, y_row, largura_util, altura_linha, fill=fill)
            x = margem_x
            for col_i, valor in enumerate(linha):
                largura = larguras[col_i]
                if col_i > 0:
                    add("q")
                    add("0.83 0.86 0.90 RG")
                    add("0.7 w")
                    add("%d %d %d %d re S" % (x, y_row, largura, altura_linha))
                    add("Q")
                txt = _pdf_fit_text(valor, max(8, int(largura / 4.6)))
                if col_i in (0, 1, 2):
                    texto(x + 3, y_row + 6, txt, tamanho=7, bold=False)
                else:
                    texto(x + 3, y_row + 6, txt, tamanho=7, bold=False)
                x += largura

        # Footer
        add("q")
        add("0.12 0.27 0.41 rg")
        add("%d %d %d 1 re f" % (margem_x, 18, largura_util))
        add("Q")
        texto(margem_x + 10, 20, "Gerado em %s" % datetime.now().strftime("%d/%m/%Y %H:%M"), tamanho=7, bold=False)
        texto(margem_x + largura_util - 110, 20, "Pagina %d" % (indice + 1), tamanho=7, bold=True)

        stream = "\n".join(comandos).encode("utf-8")
        conteudo = b"<< /Length %d >>\nstream\n" % len(stream) + stream + b"\nendstream"
        objetos.append((num_conteudo, conteudo))

        pagina_obj = (
            "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 %d %d] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents %d 0 R >>"
            % (largura_pagina, altura_pagina, num_conteudo)
        ).encode("utf-8")
        objetos.append((num_pagina, pagina_obj))

    objetos_ordenados = [
        (1, b"<< /Type /Catalog /Pages 2 0 R >>"),
        (2, ("<< /Type /Pages /Kids [%s] /Count %d >>" % (" ".join("%d 0 R" % n for n in pagina_objetos), len(pagina_objetos))).encode("utf-8")),
        (3, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"),
        (4, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>"),
    ]
    objetos_ordenados.extend(sorted(objetos, key=lambda item: item[0]))

    saida = [b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"]
    offsets = []
    pos = len(saida[0])
    for numero, conteudo in objetos_ordenados:
        bloco = ("%d 0 obj\n" % numero).encode("utf-8") + conteudo + b"\nendobj\n"
        offsets.append(pos)
        saida.append(bloco)
        pos += len(bloco)

    xref_pos = pos
    saida.append(("xref\n0 %d\n" % (len(objetos_ordenados) + 1)).encode("utf-8"))
    saida.append(b"0000000000 65535 f \n")
    for offset in offsets:
        saida.append(("%010d 00000 n \n" % offset).encode("utf-8"))
    saida.append(
        ("trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF\n" % (len(objetos_ordenados) + 1, xref_pos)).encode("utf-8")
    )

    with open(caminho, "wb") as f:
        for parte in saida:
            f.write(parte)


def montar_aba_resumo(wb, resultados, periodo_txt, company_name, logo_path, nota_vazio=None):
    ws = wb.active
    ws.title = "Resumo"

    colunas = [
        "Matricula", "Nome", "Departamento",
        "Horas Normais", "Horas Trabalhadas",
        "Atrasos (qtd)", "Atrasos (s)",
        "Saidas Antec. (qtd)", "Saidas Antec. (s)",
        "H.Extra Padrao", "H.Extra Realizada",
        "Dias Trabalhados", "Faltas",
    ]

    linha = escrever_cabecalho_pagina(ws, "Relatorio de Ponto - Resumo", periodo_txt, len(colunas), company_name, logo_path)
    escrever_linha_cabecalho_tabela(ws, linha, colunas)
    linha += 1
    linha_dados_inicio = linha
    linha_cabecalho = linha_dados_inicio - 1
    ultima_linha = linha - 1

    for r in resultados:
        valores = [
            r.get("enrollid", ""), r.get("name", ""), r.get("department", ""),
            r.get("atten_hour", ""), r.get("work_hour", ""),
            r.get("late_times", ""), _minutos_para_segundos(r.get("late_minute", "")),
            r.get("leave_times", ""), _minutos_para_segundos(r.get("leave_minute", "")),
            r.get("std_ot_hour", ""), r.get("ot_hour", ""),
            r.get("work_days", ""), r.get("absent_days", ""),
        ]
        for col_i, valor in enumerate(valores, start=1):
            cel = ws.cell(row=linha, column=col_i, value=valor)
            cel.border = BORDA_PADRAO
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.fill = PatternFill(
                start_color=COR_FAIXA if (linha - linha_dados_inicio) % 2 == 0 else COR_FAIXA_ALT,
                end_color=COR_FAIXA if (linha - linha_dados_inicio) % 2 == 0 else COR_FAIXA_ALT,
                fill_type="solid",
            )
            if col_i in (7, 9):
                cel.value = _segundos_para_duracao_excel(cel.value)
                _formatar_duracao_excel(cel)
        linha += 1
        ultima_linha = linha - 1

    if not resultados and nota_vazio:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(colunas))
        cel = ws.cell(row=linha, column=1, value=nota_vazio)
        cel.alignment = Alignment(horizontal="center")
        cel.font = Font(italic=True, color="666666")
        ultima_linha = linha

    for col_i in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = [12, 24, 22, 14, 16, 14, 14, 16, 14, 16, 16, 14, 12][col_i - 1]

    ws.freeze_panes = ws["A{}".format(max(2, linha_dados_inicio))]
    if ultima_linha >= linha_dados_inicio:
        ws.auto_filter.ref = "A{}:{}{}".format(linha_cabecalho, get_column_letter(len(colunas)), ultima_linha)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COR_TITULO


def montar_aba_detalhado(wb, resultados, periodo_txt, company_name, logo_path, nota_vazio=None):
    ws = wb.create_sheet("Detalhado")

    colunas = [
        "Data", "Dia da Semana",
        "Entrada 1", "Saida 1", "Entrada 2", "Saida 2", "Entrada 3", "Saida 3",
        "H.Normais Dia", "H.Reais Dia", "H.Extra Dia", "Pausa (min)", "Atraso (s)", "Saida Antec. (s)",
    ]

    linha = escrever_cabecalho_pagina(ws, "Relatorio de Ponto - Detalhado", periodo_txt, len(colunas), company_name, logo_path)

    for r in resultados:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(colunas))
        cel = ws.cell(row=linha, column=1)
        cel.value = "{}  ({})  -  {}".format(r.get("name", ""), r.get("enrollid", ""), r.get("department", ""))
        cel.font = Font(bold=True, size=11, color="FFFFFF")
        cel.fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
        cel.alignment = Alignment(horizontal="left", vertical="center")
        linha += 1

        escrever_linha_cabecalho_tabela(ws, linha, colunas)
        linha += 1
        linha_dados_inicio = linha

        for dia in r.get("days", []):
            horarios = list(dia.get("time", []))
            while len(horarios) < 6:
                horarios.append("")
            horarios = horarios[:6]

            intervalo_real = ""
            if horarios[1] and horarios[2]:
                intervalo_real = _intervalo_minutos(horarios[1], horarios[2])

            valores = [dia.get("day", ""), dia.get("week", "")] + horarios + [
                dia.get("fatten", ""), dia.get("fact_atten", ""), dia.get("dayot", ""), intervalo_real,
                _minutos_para_segundos(dia.get("daylate", "")), _minutos_para_segundos(dia.get("dayLeaveearly", "")),
            ]
            for col_i, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col_i, value=valor)
                cel.border = BORDA_PADRAO
                cel.alignment = Alignment(horizontal="center", vertical="center")
                cel.fill = PatternFill(
                    start_color=COR_FAIXA if (linha - linha_dados_inicio) % 2 == 0 else COR_FAIXA_ALT,
                    end_color=COR_FAIXA if (linha - linha_dados_inicio) % 2 == 0 else COR_FAIXA_ALT,
                    fill_type="solid",
                )
                if col_i in (13, 14):
                    cel.value = _segundos_para_duracao_excel(cel.value)
                    _formatar_duracao_excel(cel)
            linha += 1

        linha += 2

    if not resultados and nota_vazio:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(colunas))
        cel = ws.cell(row=linha, column=1, value=nota_vazio)
        cel.font = Font(italic=True, color="666666")
        cel.alignment = Alignment(horizontal="center", vertical="center")

    for col_i in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = [14, 14, 12, 12, 12, 12, 12, 12, 14, 14, 14, 12, 12, 14][col_i - 1]

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COR_TITULO


def gerar_planilha(resultados, inicio, fim, nome_periodo, output_dir, company_name, logo_path, vendor_name="GRB Tecnologia", nota_vazio=None):
    _garantir_diretorio(output_dir)
    periodo_txt = "{} a {}".format(inicio.strftime("%d/%m/%Y"), fim.strftime("%d/%m/%Y"))

    wb = Workbook()
    montar_aba_resumo(wb, resultados, periodo_txt, company_name, logo_path, nota_vazio=nota_vazio)
    montar_aba_detalhado(wb, resultados, periodo_txt, company_name, logo_path, nota_vazio=nota_vazio)

    nome_arquivo = "Relatorio_Ponto_{}_{}_a_{}.xlsx".format(
        nome_periodo, inicio.strftime("%Y-%m-%d"), fim.strftime("%Y-%m-%d"),
    )
    caminho_completo = os.path.join(output_dir, nome_arquivo)
    wb.save(caminho_completo)

    base_sem_extensao = os.path.splitext(caminho_completo)[0]
    cabecalhos_csv = [
        "Matricula", "Nome", "Departamento",
        "Horas Normais", "Horas Trabalhadas",
        "Atrasos (qtd)", "Atrasos (s)",
        "Saidas Antec. (qtd)", "Saidas Antec. (s)",
        "H.Extra Padrao", "H.Extra Realizada",
        "Dias Trabalhados", "Faltas",
    ]
    linhas_csv = []
    for r in resultados:
        linhas_csv.append([
            r.get("enrollid", ""), r.get("name", ""), r.get("department", ""),
            r.get("atten_hour", ""), r.get("work_hour", ""),
            r.get("late_times", ""), _minutos_para_segundos(r.get("late_minute", "")),
            r.get("leave_times", ""), _minutos_para_segundos(r.get("leave_minute", "")),
            r.get("std_ot_hour", ""), r.get("ot_hour", ""),
            r.get("work_days", ""), r.get("absent_days", ""),
        ])

    _salvar_csv(base_sem_extensao + ".csv", cabecalhos_csv, linhas_csv)
    _gerar_pdf_texto(
        base_sem_extensao + ".pdf",
        "Relatorio de Ponto - Resumo",
        periodo_txt,
        cabecalhos_csv,
        linhas_csv,
    )
    return caminho_completo, nome_arquivo


def gerar_relatorio_completo(config, inicio, fim, nome_periodo, progress_cb=None, permitir_vazio=False, nota_vazio=None, filtros=None):
    """Faz todo o fluxo: busca matriculas, busca os registros de cada uma,
    monta a planilha. progress_cb(feito, total) e chamado a cada funcionario,
    se for passado (usado pra barra de progresso na web)."""
    filtros = filtros or {}
    inicio, fim = _resolver_periodo_relatorio(inicio, fim, filtros)

    resultados = _construir_resultados_de_getlogs(config, inicio, fim, progress_cb=progress_cb)
    if not resultados:
        resultados = _construir_resultados_de_historico_local(config, inicio, fim, progress_cb=progress_cb)
    resultados = filtrar_resultados_relatorio(resultados, filtros)
    if not resultados:
        if permitir_vazio:
            nota_vazio = nota_vazio or "Nenhuma marcacao encontrada no periodo."
        else:
            if any(str(v or "").strip() for v in filtros.values()):
                raise DeviceError("Nao encontrei funcionarios para os filtros informados.")
            raise DeviceError("O leitor nao encontrou logs no periodo escolhido. Tente gerar com o historico local ou verifique se o monitor esta rodando.")

    logo_path = config.get("logo_path") or ""
    caminho, nome_arquivo = gerar_planilha(
        resultados, inicio, fim, nome_periodo,
        config["output_dir"], config["company_name"], logo_path, config.get("vendor_name", "GRB Tecnologia"),
        nota_vazio=nota_vazio,
    )
    return caminho, nome_arquivo, len(resultados)
