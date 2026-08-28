import os
import unittest

from openpyxl import load_workbook

import ponto_core


def _sample_resultados():
    return [
        {
            "enrollid": "12345",
            "name": "Alice Silva",
            "department": "TI",
            "shift_name": "Manha",
            "days": [
                {
                    "day": "2026-08-05",
                    "week": "QUA",
                    "time": ["08:00", "12:00", "13:00", "17:00"],
                    "fatten": "08:00",
                    "fact_atten": "08:00",
                    "dayot": "00:00",
                    "daylate": "0",
                    "dayLeaveearly": "0",
                }
            ],
            "atten_hour": "08:00",
            "work_hour": "08:00",
            "late_times": 0,
            "late_minute": 0,
            "leave_times": 0,
            "leave_minute": 0,
            "std_ot_hour": "08:00",
            "ot_hour": "00:00",
            "work_days": 1,
            "absent_days": 0,
        },
        {
            "enrollid": "54321",
            "name": "Bruno Costa",
            "department": "Financeiro",
            "shift_name": "Tarde",
            "days": [],
            "atten_hour": "08:00",
            "work_hour": "07:30",
            "late_times": 1,
            "late_minute": 10,
            "leave_times": 0,
            "leave_minute": 0,
            "std_ot_hour": "08:00",
            "ot_hour": "00:00",
            "work_days": 0,
            "absent_days": 1,
        },
    ]


def _sample_relatorio_resultado():
    return [
        {
            "enrollid": "12345",
            "name": "Alice Silva",
            "department": "TI",
            "shift_name": "Manha",
            "cpf": "000.000.000-00",
            "admissao": "2026-08-01",
            "atten_hour": "08:00",
            "work_hour": "08:00",
            "late_times": 0,
            "late_minute": 0,
            "leave_times": 0,
            "leave_minute": 0,
            "std_ot_hour": "08:00",
            "ot_hour": "00:00",
            "work_days": 1,
            "absent_days": 0,
            "linhas": [["05/08/2026", "08:00 12:00 13:00 17:00", "08:00 12:00 13:00 17:00", "09:00", "Manha", "OK"]],
        }
    ]


class PontoCoreFiltersTests(unittest.TestCase):
    def test_filtrar_resultados_relatorio_combina_multiplos_filtros(self):
        filtrados = ponto_core.filtrar_resultados_relatorio(
            _sample_resultados(),
            {"matricula": "12345", "nome": "alice", "departamento": "ti", "turno": "manha"},
        )
        self.assertEqual(len(filtrados), 1)
        self.assertEqual(filtrados[0]["name"], "Alice Silva")

    def test_filtrar_resultados_relatorio_retorna_vazio_para_inexistente(self):
        filtrados = ponto_core.filtrar_resultados_relatorio(
            _sample_resultados(),
            {"matricula": "99999", "nome": "Inexistente"},
        )
        self.assertEqual(filtrados, [])

    def test_resolver_periodo_relatorio_intersecta_datas(self):
        inicio, fim = ponto_core._resolver_periodo_relatorio(  # noqa: SLF001
            ponto_core.date(2026, 8, 1),
            ponto_core.date(2026, 8, 31),
            {"periodo_inicio": "2026-08-05", "periodo_fim": "2026-08-10"},
        )
        self.assertEqual(inicio.isoformat(), "2026-08-05")
        self.assertEqual(fim.isoformat(), "2026-08-10")

    def test_resolver_periodo_relatorio_valida_datas(self):
        with self.assertRaises(ValueError):
            ponto_core._resolver_periodo_relatorio(  # noqa: SLF001
                ponto_core.date(2026, 8, 1),
                ponto_core.date(2026, 8, 31),
                {"periodo_inicio": "2026-08-20", "periodo_fim": "2026-08-10"},
            )

    def test_gerar_relatorio_completo_aplica_filtros(self):
        capturado = {}

        original_getlogs = ponto_core._construir_resultados_de_getlogs
        original_local = ponto_core._construir_resultados_de_historico_local
        original_planilha = ponto_core.gerar_planilha

        try:
            ponto_core._construir_resultados_de_getlogs = lambda *args, **kwargs: _sample_resultados()  # noqa: SLF001
            ponto_core._construir_resultados_de_historico_local = lambda *args, **kwargs: []  # noqa: SLF001

            def fake_gerar_planilha(resultados, inicio, fim, nome_periodo, output_dir, company_name, logo_path, vendor_name="GRB Tecnologia", nota_vazio=None):
                capturado["resultados"] = resultados
                capturado["inicio"] = inicio
                capturado["fim"] = fim
                capturado["nome_periodo"] = nome_periodo
                return os.path.join(output_dir, "relatorio.xlsx"), "relatorio.xlsx"

            ponto_core.gerar_planilha = fake_gerar_planilha

            caminho, nome_arquivo, qtd = ponto_core.gerar_relatorio_completo(
                {"output_dir": os.getcwd(), "company_name": "Teste", "logo_path": ""},
                ponto_core.date(2026, 8, 1),
                ponto_core.date(2026, 8, 31),
                "Mensal",
                filtros={"matricula": "12345", "departamento": "TI", "periodo_inicio": "2026-08-05", "periodo_fim": "2026-08-10"},
            )
        finally:
            ponto_core._construir_resultados_de_getlogs = original_getlogs  # noqa: SLF001
            ponto_core._construir_resultados_de_historico_local = original_local  # noqa: SLF001
            ponto_core.gerar_planilha = original_planilha

        self.assertTrue(caminho.endswith("relatorio.xlsx"))
        self.assertEqual(nome_arquivo, "relatorio.xlsx")
        self.assertEqual(qtd, 1)
        self.assertEqual(capturado["inicio"].isoformat(), "2026-08-05")
        self.assertEqual(capturado["fim"].isoformat(), "2026-08-10")
        self.assertEqual(len(capturado["resultados"]), 1)
        self.assertEqual(capturado["resultados"][0]["name"], "Alice Silva")

    def test_gerar_relatorio_completo_lanca_erro_com_filtro_sem_resultado(self):
        original_getlogs = ponto_core._construir_resultados_de_getlogs
        original_local = ponto_core._construir_resultados_de_historico_local
        original_planilha = ponto_core.gerar_planilha

        try:
            ponto_core._construir_resultados_de_getlogs = lambda *args, **kwargs: _sample_resultados()  # noqa: SLF001
            ponto_core._construir_resultados_de_historico_local = lambda *args, **kwargs: []  # noqa: SLF001
            ponto_core.gerar_planilha = lambda *args, **kwargs: None

            with self.assertRaises(ponto_core.DeviceError):
                ponto_core.gerar_relatorio_completo(
                    {"output_dir": os.getcwd(), "company_name": "Teste", "logo_path": ""},
                    ponto_core.date(2026, 8, 1),
                    ponto_core.date(2026, 8, 31),
                    "Mensal",
                    filtros={"matricula": "99999"},
                )
        finally:
            ponto_core._construir_resultados_de_getlogs = original_getlogs  # noqa: SLF001
            ponto_core._construir_resultados_de_historico_local = original_local  # noqa: SLF001
            ponto_core.gerar_planilha = original_planilha

    def test_gerar_planilha_cria_xlsx_csv_pdf(self):
        caminho, nome_arquivo = ponto_core.gerar_planilha(
            _sample_relatorio_resultado(),
            ponto_core.date(2026, 8, 1),
            ponto_core.date(2026, 8, 31),
            "Mensal",
            os.getcwd(),
            "Empresa Teste",
            "",
        )

        base = os.path.splitext(caminho)[0]
        self.assertTrue(nome_arquivo.endswith(".xlsx"))
        self.assertTrue(os.path.isfile(caminho))
        self.assertTrue(os.path.isfile(base + ".csv"))
        self.assertTrue(os.path.isfile(base + ".pdf"))

        wb = load_workbook(caminho)
        self.assertIn("Resumo", wb.sheetnames)
        self.assertIn("Detalhado", wb.sheetnames)
        self.assertEqual(wb["Resumo"]["A1"].value, "Empresa Teste  •  GRB Tecnologia")

        with open(base + ".csv", "r", encoding="utf-8-sig") as f_csv:
            conteudo_csv = f_csv.read()
        self.assertIn("Alice Silva", conteudo_csv)

        with open(base + ".pdf", "rb") as f_pdf:
            self.assertEqual(f_pdf.read(5), b"%PDF-")


if __name__ == "__main__":
    unittest.main()
